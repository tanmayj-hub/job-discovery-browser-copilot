from __future__ import annotations

import json
from pathlib import Path

import yaml
from openpyxl import Workbook

from onboarding.live_discovery import FetchedPage
from onboarding.source_onboarding import (
    apply_approved_candidates,
    audit_large_company_list,
    generate_candidates,
    generate_candidates_from_input,
    load_candidate_file,
    refresh_sources,
    weekly_source_check,
)
from storage.db import (
    create_intervention,
    initialize_database,
    record_source_observation,
    upsert_companies,
)


def _write_companies_yaml(path: Path, companies: list[dict[str, object]]) -> None:
    path.write_text(
        yaml.safe_dump({"companies": companies}, sort_keys=False),
        encoding="utf-8",
    )


def _write_starter_yaml(path: Path, companies: list[dict[str, object]]) -> None:
    path.write_text(
        yaml.safe_dump({"companies": companies}, sort_keys=False),
        encoding="utf-8",
    )


def _write_reference_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Companies"
    sheet.append(
        [
            "Company",
            "Careers page URL (fill in)",
            "website category",
            "Sector",
            "Category",
            "Canada hubs / notes",
            "Role families",
            "Suggested search keywords",
            "Early-career pipeline",
            "Priority",
            "Monitoring hint",
            "Status",
            "Last checked",
            "Notes",
        ]
    )
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _write_large_list_workbook(
    path: Path,
    rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Companies"
    headers = [
        "Company",
        "Careers page URL (fill in)",
        "website category",
        "Sector",
        "Category",
        "Canada hubs / notes",
        "Role families",
        "Suggested search keywords",
        "Early-career pipeline",
        "Priority",
        "Monitoring hint",
        "Status",
        "Last checked",
        "Notes",
    ]
    sheet.append(headers)
    for row_index, row in enumerate(rows, start=2):
        values = [
            row.get("company_name"),
            row.get("career_display_text"),
            row.get("website_category"),
            row.get("sector"),
            row.get("category"),
            row.get("canada_hubs_notes"),
            row.get("role_families"),
            row.get("keywords"),
            row.get("early_career_pipeline"),
            row.get("priority"),
            row.get("monitoring_hint"),
            row.get("status"),
            row.get("last_checked"),
            row.get("notes"),
        ]
        sheet.append(values)
        hyperlink = row.get("career_hyperlink")
        if hyperlink:
            sheet.cell(row=row_index, column=2).hyperlink = str(hyperlink)
    workbook.save(path)


def _fake_fetcher(
    mapping: dict[str, tuple[str, str, int] | Exception],
    counter: list[str] | None = None,
):
    def fetch(url: str, _timeout: int) -> FetchedPage:
        normalized = url.rstrip("/")
        if counter is not None:
            counter.append(normalized)
        value = mapping.get(normalized)
        if value is None:
            return FetchedPage(
                requested_url=url,
                final_url=url,
                status_code=200,
                text="<html></html>",
                content_type="text/html; charset=utf-8",
            )
        if isinstance(value, Exception):
            raise value
        final_url, text, status_code = value
        return FetchedPage(
            requested_url=url,
            final_url=final_url,
            status_code=status_code,
            text=text,
            content_type="text/html; charset=utf-8",
        )

    return fetch


def _allow_all(_url: str) -> bool:
    return True


def _sample_company(**overrides: object) -> dict[str, object]:
    company = {
        "name": "Example Co",
        "sector": "IT Consulting & Systems Integrators",
        "category": "Consulting/SI",
        "careers_url": "https://careers.example.com/jobs",
        "website_category": "company-careers",
        "ats_hint": "",
        "canada_hubs_notes": "Canada",
        "role_families": ["Cloud"],
        "keywords": ["cloud"],
        "priority": "High",
        "monitoring_hint": "Monitor",
        "status": "Watching",
        "source_mode": "browser_allowed",
    }
    company.update(overrides)
    return company


def test_existing_configured_company_produces_high_confidence_candidate(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter_career_urls.yaml"
    _write_companies_yaml(
        companies_path,
        [
            {
                "name": "Example Bank",
                "sector": "Banking & Capital Markets",
                "category": "Bank/Market",
                "careers_url": "https://jobs.examplebank.com/roles",
                "website_category": "greenhouse",
                "ats_hint": "greenhouse",
                "source_mode": "api_allowed",
            }
        ],
    )
    _write_starter_yaml(starter_path, [])

    candidates = generate_candidates(
        ["Example Bank"],
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=(),
    )

    assert len(candidates) == 1
    assert candidates[0].reason == "existing_config_match"
    assert candidates[0].confidence == "high"
    assert candidates[0].detected_ats_type == "greenhouse"
    assert candidates[0].suggested_source_mode == "api_allowed"
    assert candidates[0].approved is False


def test_starter_career_url_produces_reviewable_candidate(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter_career_urls.yaml"
    _write_companies_yaml(companies_path, [])
    _write_starter_yaml(
        starter_path,
        [
            {
                "name": "Starter Co",
                "careers_url": "https://www.starterco.com/careers",
                "confidence": "medium",
                "notes": "Verified starter URL",
            }
        ],
    )

    candidates = generate_candidates(
        ["Starter Co"],
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=(),
    )

    assert candidates[0].reason == "starter_career_url_match"
    assert candidates[0].confidence == "medium"
    assert candidates[0].needs_review is True
    assert candidates[0].candidate_careers_url == "https://www.starterco.com/careers"


def test_missing_url_produces_low_confidence_candidate(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter_career_urls.yaml"
    _write_companies_yaml(companies_path, [])
    _write_starter_yaml(starter_path, [])

    candidates = generate_candidates(
        ["Unknown Co"],
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=(),
    )

    assert candidates[0].confidence == "low"
    assert candidates[0].needs_review is True
    assert candidates[0].reason == "missing_candidate_url"
    assert candidates[0].suggested_source_mode == "needs_url"


def test_greenhouse_candidate_detects_api_allowed(tmp_path: Path) -> None:
    reference_path = tmp_path / "reference.xlsx"
    _write_reference_workbook(
        reference_path,
        [
            [
                "Greenhouse Co",
                "https://boards.greenhouse.io/example",
                "greenhouse",
                "IT Consulting & Systems Integrators",
                "Consulting/SI",
                "Canada",
                "Cloud",
                "cloud",
                "",
                "High",
                "Monitor",
                "Watching",
                "",
                "",
            ]
        ],
    )

    candidate = generate_candidates(
        ["Greenhouse Co"],
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(reference_path,),
    )[0]

    assert candidate.detected_ats_type == "greenhouse"
    assert candidate.suggested_source_mode == "api_allowed"
    assert candidate.confidence == "high"


def test_lever_candidate_detects_api_allowed(tmp_path: Path) -> None:
    reference_path = tmp_path / "reference.xlsx"
    _write_reference_workbook(
        reference_path,
        [
            [
                "Lever Co",
                "https://jobs.lever.co/example",
                "lever",
                "IT Consulting & Systems Integrators",
                "Consulting/SI",
                "Canada",
                "Cloud",
                "cloud",
                "",
                "High",
                "Monitor",
                "Watching",
                "",
                "",
            ]
        ],
    )

    candidate = generate_candidates(
        ["Lever Co"],
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(reference_path,),
    )[0]

    assert candidate.detected_ats_type == "lever"
    assert candidate.suggested_source_mode == "api_allowed"
    assert candidate.confidence == "high"


def test_ashby_candidate_detects_api_allowed(tmp_path: Path) -> None:
    reference_path = tmp_path / "reference.xlsx"
    _write_reference_workbook(
        reference_path,
        [
            [
                "Ashby Co",
                "https://jobs.ashbyhq.com/example",
                "ashby",
                "IT Consulting & Systems Integrators",
                "Consulting/SI",
                "Canada",
                "Cloud",
                "cloud",
                "",
                "High",
                "Monitor",
                "Watching",
                "",
                "",
            ]
        ],
    )

    candidate = generate_candidates(
        ["Ashby Co"],
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(reference_path,),
    )[0]

    assert candidate.detected_ats_type == "ashby"
    assert candidate.suggested_source_mode == "api_allowed"
    assert candidate.confidence == "high"


def test_workday_candidate_suggests_human_in_loop(tmp_path: Path) -> None:
    reference_path = tmp_path / "reference.xlsx"
    _write_reference_workbook(
        reference_path,
        [
            [
                "Workday Co",
                "https://example.myworkdayjobs.com/en-US/careers",
                "workday",
                "IT Consulting & Systems Integrators",
                "Consulting/SI",
                "Canada",
                "Cloud",
                "cloud",
                "",
                "High",
                "Monitor",
                "Watching",
                "",
                "",
            ]
        ],
    )

    candidate = generate_candidates(
        ["Workday Co"],
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(reference_path,),
    )[0]

    assert candidate.detected_ats_type == "workday"
    assert candidate.suggested_source_mode == "human_in_loop"


def test_restricted_board_candidate_is_manual_only_and_not_auto_applied(
    tmp_path: Path,
) -> None:
    reference_path = tmp_path / "reference.xlsx"
    companies_path = tmp_path / "companies.yaml"
    candidates_path = tmp_path / "candidates.yaml"
    _write_companies_yaml(companies_path, [])
    _write_reference_workbook(
        reference_path,
        [
            [
                "Restricted Co",
                "https://www.linkedin.com/jobs/view/123",
                "LinkedIn",
                "IT Consulting & Systems Integrators",
                "Consulting/SI",
                "Canada",
                "Cloud",
                "cloud",
                "",
                "High",
                "Monitor",
                "Watching",
                "",
                "",
            ]
        ],
    )

    generated = generate_candidates(
        ["Restricted Co"],
        companies_path=companies_path,
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(reference_path,),
    )
    generated[0].approved = True
    candidates_path.write_text(
        yaml.safe_dump(
            {"candidates": [generated[0].model_dump(exclude_none=True)]},
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    summary = apply_approved_candidates(
        input_path=candidates_path,
        companies_path=companies_path,
    )
    payload = yaml.safe_load(companies_path.read_text(encoding="utf-8"))

    assert generated[0].suggested_source_mode == "manual_only"
    assert generated[0].detected_ats_type == "restricted_board"
    assert summary["applied"] == 0
    assert summary["skipped_restricted"] == 1
    assert payload["companies"] == []


def test_candidate_output_file_is_created(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.txt"
    output_path = tmp_path / "candidates.yaml"
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter_career_urls.yaml"
    input_path.write_text("Example Co\nUnknown Co\n", encoding="utf-8")
    _write_companies_yaml(
        companies_path,
        [
            {
                "name": "Example Co",
                "sector": "Banking & Capital Markets",
                "category": "Bank/Market",
                "careers_url": "https://www.example.com/careers",
                "source_mode": "browser_allowed",
            }
        ],
    )
    _write_starter_yaml(starter_path, [])

    generate_candidates_from_input(
        input_path=input_path,
        output_path=output_path,
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=(),
    )

    payload = yaml.safe_load(output_path.read_text(encoding="utf-8"))
    assert output_path.exists()
    assert len(payload["candidates"]) == 2


def test_apply_command_only_applies_approved_candidates(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    candidates_path = tmp_path / "candidates.yaml"
    _write_companies_yaml(companies_path, [])
    candidates_path.write_text(
        yaml.safe_dump(
            {
                "candidates": [
                    {
                        "company_name": "Approved Co",
                        "candidate_careers_url": "https://approved.example.com/careers",
                        "detected_ats_type": None,
                        "suggested_source_mode": "browser_allowed",
                        "confidence": "medium",
                        "needs_review": True,
                        "reason": "starter_career_url_match",
                        "evidence": ["starter"],
                        "approved": True,
                        "sector": "IT Consulting & Systems Integrators",
                        "category": "Consulting/SI",
                        "status": "Watching",
                        "role_families": ["Cloud"],
                        "keywords": ["cloud"],
                    },
                    {
                        "company_name": "Pending Co",
                        "candidate_careers_url": "https://pending.example.com/careers",
                        "detected_ats_type": None,
                        "suggested_source_mode": "browser_allowed",
                        "confidence": "medium",
                        "needs_review": True,
                        "reason": "starter_career_url_match",
                        "evidence": ["starter"],
                        "approved": False,
                        "sector": "IT Consulting & Systems Integrators",
                        "category": "Consulting/SI",
                        "status": "Watching",
                        "role_families": ["Cloud"],
                        "keywords": ["cloud"],
                    },
                ]
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    summary = apply_approved_candidates(
        input_path=candidates_path,
        companies_path=companies_path,
    )
    payload = yaml.safe_load(companies_path.read_text(encoding="utf-8"))

    assert summary["applied"] == 1
    assert summary["skipped_unapproved"] == 1
    assert payload["companies"][0]["name"] == "Approved Co"
    assert payload["companies"][0]["source_mode"] == "browser_allowed"


def test_apply_does_not_overwrite_existing_companies_by_default(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    candidates_path = tmp_path / "candidates.yaml"
    _write_companies_yaml(
        companies_path,
        [
            {
                "name": "Existing Co",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
                "careers_url": "https://old.example.com/careers",
                "source_mode": "browser_allowed",
            }
        ],
    )
    candidates_path.write_text(
        yaml.safe_dump(
            {
                "candidates": [
                    {
                        "company_name": "Existing Co",
                        "candidate_careers_url": "https://new.example.com/careers",
                        "detected_ats_type": "workday",
                        "suggested_source_mode": "human_in_loop",
                        "confidence": "high",
                        "needs_review": False,
                        "reason": "existing_config_match",
                        "evidence": ["config"],
                        "approved": True,
                        "sector": "IT Consulting & Systems Integrators",
                        "category": "Consulting/SI",
                        "status": "Watching",
                        "role_families": ["Cloud"],
                        "keywords": ["cloud"],
                    }
                ]
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    summary = apply_approved_candidates(
        input_path=candidates_path,
        companies_path=companies_path,
    )
    payload = yaml.safe_load(companies_path.read_text(encoding="utf-8"))

    assert summary["applied"] == 0
    assert summary["skipped_existing"] == 1
    assert payload["companies"][0]["careers_url"] == "https://old.example.com/careers"
    assert payload["companies"][0]["source_mode"] == "browser_allowed"


def test_load_candidate_file_reads_generated_yaml(tmp_path: Path) -> None:
    candidates_path = tmp_path / "candidates.yaml"
    candidates_path.write_text(
        yaml.safe_dump(
            {
                "candidates": [
                    {
                        "company_name": "Example Co",
                        "candidate_careers_url": "https://example.com/careers",
                        "suggested_source_mode": "browser_allowed",
                        "confidence": "medium",
                        "needs_review": True,
                        "reason": "starter_career_url_match",
                        "evidence": ["starter"],
                        "approved": False,
                    }
                ]
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    candidates = load_candidate_file(candidates_path)

    assert len(candidates) == 1
    assert candidates[0].company_name == "Example Co"


def test_without_live_discovery_behavior_remains_internal_only(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter.yaml"
    input_path.write_text(
        "company_name,website_url\nExample Co,https://www.example.com\n",
        encoding="utf-8",
    )
    _write_companies_yaml(companies_path, [])
    _write_starter_yaml(starter_path, [])

    calls: list[str] = []
    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=(),
        live_discovery=False,
        fetch_page=_fake_fetcher({}, calls),
        robots_allowed=_allow_all,
    )

    assert len(candidates) == 1
    assert calls == []
    assert candidates[0].reason == "missing_candidate_url"


def test_with_live_discovery_and_provided_website_discovers_homepage_careers_link(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nExample Co,https://www.example.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.example.com": (
                "https://www.example.com",
                '<a href="/careers">Careers</a>',
                200,
            ),
            "https://www.example.com/careers": (
                "https://www.example.com/careers",
                "<html><body>Open roles</body></html>",
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    live_candidates = [
        item for item in candidates if item.reason == "live_discovery_careers_link"
    ]
    assert live_candidates
    assert live_candidates[0].candidate_careers_url == "https://www.example.com/careers"


def test_same_domain_careers_url_becomes_browser_allowed_with_medium_confidence(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nBrowser Co,https://www.browserco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.browserco.com": (
                "https://www.browserco.com",
                '<a href="/jobs">Jobs</a>',
                200,
            ),
            "https://www.browserco.com/jobs": (
                "https://www.browserco.com/jobs",
                "<html><body>Roles</body></html>",
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    careers_candidate = next(
        item for item in candidates if item.candidate_careers_url == "https://www.browserco.com/jobs"
    )
    assert careers_candidate.suggested_source_mode == "browser_allowed"
    assert careers_candidate.confidence == "medium"


def test_official_page_linking_to_greenhouse_produces_high_confidence_candidate(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nGreenhouse Co,https://www.greenhouseco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.greenhouseco.com": (
                "https://www.greenhouseco.com",
                '<a href="https://boards.greenhouse.io/example">Open Roles</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    ats_candidate = next(item for item in candidates if item.detected_ats_type == "greenhouse")
    assert ats_candidate.confidence == "high"
    assert ats_candidate.suggested_source_mode == "api_allowed"


def test_official_page_linking_to_lever_produces_high_confidence_candidate(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nLever Co,https://www.leverco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.leverco.com": (
                "https://www.leverco.com",
                '<a href="https://jobs.lever.co/example">Open Roles</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    ats_candidate = next(item for item in candidates if item.detected_ats_type == "lever")
    assert ats_candidate.confidence == "high"
    assert ats_candidate.suggested_source_mode == "api_allowed"


def test_official_page_linking_to_ashby_produces_high_confidence_candidate(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nAshby Co,https://www.ashbyco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.ashbyco.com": (
                "https://www.ashbyco.com",
                '<a href="https://jobs.ashbyhq.com/example">Open Roles</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    ats_candidate = next(item for item in candidates if item.detected_ats_type == "ashby")
    assert ats_candidate.confidence == "high"
    assert ats_candidate.suggested_source_mode == "api_allowed"


def test_workday_link_is_detected_and_marked_human_in_loop(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nWorkday Co,https://www.workdayco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.workdayco.com": (
                "https://www.workdayco.com",
                '<a href="https://example.myworkdayjobs.com/en-US/careers">Careers</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    workday_candidate = next(item for item in candidates if item.detected_ats_type == "workday")
    assert workday_candidate.suggested_source_mode == "human_in_loop"


def test_restricted_live_discovery_link_is_manual_only_and_not_auto_applied(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    candidates_path = tmp_path / "out.yaml"
    companies_path = tmp_path / "companies.yaml"
    input_path.write_text(
        "company_name,website_url\nRestricted Co,https://www.restrictedco.com\n",
        encoding="utf-8",
    )
    _write_companies_yaml(companies_path, [])
    fetcher = _fake_fetcher(
        {
            "https://www.restrictedco.com": (
                "https://www.restrictedco.com",
                '<a href="https://www.indeed.com/viewjob?jk=1">Jobs</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=candidates_path,
        companies_path=companies_path,
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )
    restricted = next(item for item in candidates if item.detected_ats_type == "restricted_board")
    restricted.approved = True
    candidates_path.write_text(
        yaml.safe_dump(
            {"candidates": [restricted.model_dump(exclude_none=True)]},
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    summary = apply_approved_candidates(
        input_path=candidates_path,
        companies_path=companies_path,
    )

    assert restricted.suggested_source_mode == "manual_only"
    assert summary["applied"] == 0
    assert summary["skipped_restricted"] == 1


def test_discovery_respects_crawl_page_limit(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nLimit Co,https://www.limitco.com\n",
        encoding="utf-8",
    )
    calls: list[str] = []
    fetcher = _fake_fetcher(
        {
            "https://www.limitco.com": (
                "https://www.limitco.com",
                '<a href="/careers">Careers</a><a href="/jobs">Jobs</a>',
                200,
            ),
            "https://www.limitco.com/careers": (
                "https://www.limitco.com/careers",
                '<a href="/team">Team</a>',
                200,
            ),
            "https://www.limitco.com/jobs": (
                "https://www.limitco.com/jobs",
                "<html></html>",
                200,
            ),
            "https://www.limitco.com/team": (
                "https://www.limitco.com/team",
                "<html></html>",
                200,
            ),
        },
        calls,
    )

    generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        max_pages_per_company=2,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    assert len(calls) == 2


def test_live_discovery_handles_fetch_error_with_low_confidence_candidate(
    tmp_path: Path,
) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nTimeout Co,https://www.timeoutco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher({"https://www.timeoutco.com": TimeoutError("timed out")})

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    error_candidate = next(item for item in candidates if item.reason == "live_discovery_failed")
    assert error_candidate.confidence == "low"
    assert error_candidate.needs_review is True


def test_candidate_evidence_includes_source_parent_url_and_matched_text(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nEvidence Co,https://www.evidenceco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.evidenceco.com": (
                "https://www.evidenceco.com",
                '<a href="/careers">Join Us</a>',
                200,
            ),
            "https://www.evidenceco.com/careers": (
                "https://www.evidenceco.com/careers",
                "<html></html>",
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    candidate = next(item for item in candidates if item.reason == "live_discovery_careers_link")
    evidence_text = " ".join(candidate.evidence)
    assert "source=homepage_link" in evidence_text
    assert "parent_url=https://www.evidenceco.com" in evidence_text
    assert "matched_text=Join Us" in evidence_text


def test_live_discovery_ignores_locale_switcher_variants(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nLocale Co,https://www.localeco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.localeco.com": (
                "https://www.localeco.com/en-ca",
                '<a href="/en-ca/jobs">Jobs</a>',
                200,
            ),
            "https://www.localeco.com/en-ca/jobs": (
                "https://www.localeco.com/en-ca/jobs",
                '<a href="/ae/jobs">English</a><a href="/fr-ca/jobs">Francais</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    live_urls = {
        item.candidate_careers_url
        for item in candidates
        if item.reason == "live_discovery_careers_link"
    }
    assert "https://www.localeco.com/en-ca/jobs" in live_urls
    assert "https://www.localeco.com/ae/jobs" not in live_urls
    assert "https://www.localeco.com/fr-ca/jobs" not in live_urls


def test_live_discovery_ignores_numeric_job_detail_links(tmp_path: Path) -> None:
    input_path = tmp_path / "companies.csv"
    input_path.write_text(
        "company_name,website_url\nDetail Co,https://www.detailco.com\n",
        encoding="utf-8",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.detailco.com": (
                "https://www.detailco.com",
                '<a href="/careers">Careers</a>',
                200,
            ),
            "https://www.detailco.com/careers": (
                "https://www.detailco.com/careers",
                '<a href="/careers/engineering">Engineering</a>'
                '<a href="/careers/12345">Platform Engineer</a>',
                200,
            ),
        }
    )

    candidates = generate_candidates_from_input(
        input_path=input_path,
        output_path=tmp_path / "out.yaml",
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        reference_workbooks=(),
        live_discovery=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    live_urls = {
        item.candidate_careers_url
        for item in candidates
        if item.reason == "live_discovery_careers_link"
    }
    assert "https://www.detailco.com/careers" in live_urls
    assert "https://www.detailco.com/careers/12345" not in live_urls


def test_readiness_audit_identifies_already_configured_company(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter.yaml"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "EQ Bank",
                "career_display_text": "EQ Careers",
                "sector": "Banking & Capital Markets",
                "category": "Bank/Market",
                "priority": "High",
                "status": "Watching",
            }
        ],
    )
    _write_companies_yaml(
        companies_path,
        [
            {
                "name": "EQ Bank (Equitable Bank)",
                "careers_url": "https://www.eqbank.ca/careers",
                "source_mode": "browser_allowed",
                "ats_hint": "",
                "status": "Watching",
            }
        ],
    )
    _write_starter_yaml(starter_path, [])

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=companies_path,
        starter_path=starter_path,
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )

    row = result["rows"][0]
    assert row.existing_config_match is True
    assert row.existing_config_url == "https://www.eqbank.ca/careers"
    assert row.readiness_status == "already_configured"


def test_readiness_audit_extracts_spreadsheet_hyperlink_url(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Hyperlink Co",
                "career_display_text": "Open roles",
                "career_hyperlink": "https://jobs.lever.co/hyperlink-co",
                "website_category": "lever",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
                "priority": "High",
                "status": "Watching",
            }
        ],
    )
    _write_companies_yaml(tmp_path / "companies.yaml", [])
    _write_starter_yaml(tmp_path / "starter.yaml", [])

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )

    row = result["rows"][0]
    assert row.spreadsheet_career_url_or_hyperlink == "https://jobs.lever.co/hyperlink-co"
    assert row.detected_ats_type == "lever"
    assert row.suggested_source_mode == "api_allowed"


def test_readiness_audit_distinguishes_display_text_from_hyperlink_target(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Display Co",
                "career_display_text": "Display label only",
                "career_hyperlink": "https://boards.greenhouse.io/displayco",
                "website_category": "greenhouse",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            }
        ],
    )
    _write_companies_yaml(tmp_path / "companies.yaml", [])
    _write_starter_yaml(tmp_path / "starter.yaml", [])

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )

    row = result["rows"][0]
    assert row.spreadsheet_career_display_text == "Display label only"
    assert (
        row.spreadsheet_career_url_or_hyperlink
        == "https://boards.greenhouse.io/displayco"
    )


def test_readiness_audit_marks_missing_urls_for_manual_or_website_follow_up(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Manual Co",
                "career_display_text": "Careers page coming soon",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            },
            {
                "company_name": "Website Co",
                "career_display_text": "",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            },
        ],
    )
    _write_companies_yaml(tmp_path / "companies.yaml", [])
    _write_starter_yaml(tmp_path / "starter.yaml", [])

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )

    manual_row = next(row for row in result["rows"] if row.company_name == "Manual Co")
    website_row = next(row for row in result["rows"] if row.company_name == "Website Co")
    assert manual_row.readiness_status == "needs_manual_career_url"
    assert website_row.readiness_status == "needs_website_url"


def test_readiness_audit_does_not_auto_apply_candidates(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    companies_path = tmp_path / "companies.yaml"
    starter_path = tmp_path / "starter.yaml"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Candidate Co",
                "career_display_text": "Roles",
                "career_hyperlink": "https://candidate.example.com/careers",
                "website_category": "company-careers",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            }
        ],
    )
    _write_companies_yaml(companies_path, [])
    _write_starter_yaml(starter_path, [])
    before = companies_path.read_text(encoding="utf-8")

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=companies_path,
        starter_path=starter_path,
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )
    after = companies_path.read_text(encoding="utf-8")

    assert len(result["candidates"]) == 1
    assert before == after


def test_large_list_candidate_output_defaults_approved_false(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    candidates_path = tmp_path / "candidates.yaml"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Starter Candidate Co",
                "career_display_text": "",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            }
        ],
    )
    _write_companies_yaml(tmp_path / "companies.yaml", [])
    _write_starter_yaml(
        tmp_path / "starter.yaml",
        [
            {
                "name": "Starter Candidate Co",
                "careers_url": "https://starter.example.com/careers",
                "confidence": "medium",
                "notes": "starter",
            }
        ],
    )

    audit_large_company_list(
        input_path=workbook_path,
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=candidates_path,
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )
    payload = yaml.safe_load(candidates_path.read_text(encoding="utf-8"))

    assert payload["candidates"][0]["approved"] is False


def test_restricted_large_list_candidate_is_manual_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "large-list.xlsx"
    _write_large_list_workbook(
        workbook_path,
        [
            {
                "company_name": "Restricted Large Co",
                "career_display_text": "LinkedIn jobs",
                "career_hyperlink": "https://www.linkedin.com/jobs/view/123",
                "website_category": "LinkedIn",
                "sector": "IT Consulting & Systems Integrators",
                "category": "Consulting/SI",
            }
        ],
    )
    _write_companies_yaml(tmp_path / "companies.yaml", [])
    _write_starter_yaml(tmp_path / "starter.yaml", [])

    result = audit_large_company_list(
        input_path=workbook_path,
        companies_path=tmp_path / "companies.yaml",
        starter_path=tmp_path / "starter.yaml",
        readiness_output_path=tmp_path / "readiness.csv",
        candidates_output_path=tmp_path / "candidates.yaml",
        report_output_path=tmp_path / "report.md",
        needs_website_output_path=tmp_path / "needs-website.csv",
        inspected_input_paths=(workbook_path,),
    )

    row = result["rows"][0]
    candidate = result["candidates"][0]
    assert row.readiness_status == "restricted_manual_only"
    assert candidate.suggested_source_mode == "manual_only"
    assert candidate.approved is False


def test_problem_source_refresh_discovers_replacement_candidate(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    state_path = tmp_path / "source-health-state.json"
    output_path = tmp_path / "refresh.yaml"
    _write_companies_yaml(
        companies_path,
        [
            _sample_company(
                name="Refresh Co",
                careers_url="https://www.refreshco.com/careers-old",
            )
        ],
    )
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Refresh Co", careers_url="https://www.refreshco.com/careers-old")])
    record_source_observation(
        connection,
        company_name="Refresh Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://www.refreshco.com/careers-old",
        collector="browser_after_jsonld",
        status="error",
        error="HTTP 404 not found",
    )
    record_source_observation(
        connection,
        company_name="Refresh Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://www.refreshco.com/careers-old",
        collector="browser_after_jsonld",
        status="error",
        error="HTTP 404 not found",
    )
    fetcher = _fake_fetcher(
        {
            "https://www.refreshco.com/careers-old": TimeoutError("timed out"),
            "https://www.refreshco.com": (
                "https://www.refreshco.com",
                '<a href="/careers">Careers</a>',
                200,
            ),
            "https://www.refreshco.com/careers": (
                "https://www.refreshco.com/careers",
                '<a href="https://boards.greenhouse.io/refreshco">Open Roles</a>',
                200,
            ),
        }
    )

    candidates = refresh_sources(
        output_path=output_path,
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        only_problem_sources=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    replacement = next(item for item in candidates if item.detected_ats_type == "greenhouse")
    assert replacement.company_name == "Refresh Co"
    assert replacement.current_careers_url == "https://www.refreshco.com/careers-old"
    assert replacement.suggested_action == "replace_with_candidate"


def test_refresh_candidate_output_includes_current_and_candidate_urls(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    output_path = tmp_path / "refresh.yaml"
    _write_companies_yaml(companies_path, [_sample_company(name="Refresh Co")])
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Refresh Co")])
    record_source_observation(
        connection,
        company_name="Refresh Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    record_source_observation(
        connection,
        company_name="Refresh Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    fetcher = _fake_fetcher(
        {
            "https://careers.example.com/jobs": (
                "https://careers.example.com/jobs",
                '<a href="https://jobs.lever.co/example">Apply</a>',
                200,
            ),
        }
    )

    candidates = refresh_sources(
        output_path=output_path,
        companies_path=companies_path,
        db_path=db_path,
        state_path=tmp_path / "state.json",
        only_problem_sources=True,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    candidate = next(item for item in candidates if item.detected_ats_type == "lever")
    assert candidate.current_careers_url == "https://careers.example.com/jobs"
    assert candidate.candidate_job_board_url == "https://jobs.lever.co/example"


def test_approved_replacement_only_updates_config_with_update_existing(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    candidate_path = tmp_path / "refresh.yaml"
    _write_companies_yaml(
        companies_path,
        [_sample_company(name="Replace Co", careers_url="https://old.example.com/careers")],
    )
    candidate_path.write_text(
        yaml.safe_dump(
            {
                "candidates": [
                    {
                        "company_name": "Replace Co",
                        "current_careers_url": "https://old.example.com/careers",
                        "candidate_careers_url": "https://new.example.com/careers",
                        "suggested_source_mode": "browser_allowed",
                        "confidence": "medium",
                        "needs_review": True,
                        "reason": "live_discovery_careers_link",
                        "evidence": ["source=source_refresh"],
                        "approved": True,
                        "sector": "IT Consulting & Systems Integrators",
                        "category": "Consulting/SI",
                        "status": "Watching",
                        "role_families": ["Cloud"],
                        "keywords": ["cloud"],
                    }
                ]
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    skipped_summary = apply_approved_candidates(
        input_path=candidate_path,
        companies_path=companies_path,
        update_existing=False,
    )
    skipped_payload = yaml.safe_load(companies_path.read_text(encoding="utf-8"))

    applied_summary = apply_approved_candidates(
        input_path=candidate_path,
        companies_path=companies_path,
        update_existing=True,
    )
    applied_payload = yaml.safe_load(companies_path.read_text(encoding="utf-8"))

    assert skipped_summary["applied"] == 0
    assert skipped_payload["companies"][0]["careers_url"] == "https://old.example.com/careers"
    assert applied_summary["applied"] == 1
    assert applied_payload["companies"][0]["careers_url"] == "https://new.example.com/careers"


def test_non_problem_source_is_skipped_when_only_problem_sources_is_used(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    _write_companies_yaml(companies_path, [_sample_company(name="Healthy Co")])
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Healthy Co")])
    record_source_observation(
        connection,
        company_name="Healthy Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="completed",
        jobs_discovered=3,
    )

    candidates = refresh_sources(
        output_path=tmp_path / "refresh.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=tmp_path / "state.json",
        only_problem_sources=True,
        fetch_page=_fake_fetcher({}, []),
        robots_allowed=_allow_all,
    )

    assert candidates == []


def test_weekly_check_skips_recently_checked_sources(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    state_path = tmp_path / "state.json"
    _write_companies_yaml(companies_path, [_sample_company(name="Weekly Co")])
    state_path.write_text(
        json.dumps(
            {
                "sources": {
                    "weekly co": {
                        "last_health_check_at": "2099-01-01T00:00:00Z"
                    }
                }
            }
        ),
        encoding="utf-8",
    )

    candidates = weekly_source_check(
        output_path=tmp_path / "weekly.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        fetch_page=_fake_fetcher({}, []),
        robots_allowed=_allow_all,
    )

    assert candidates == []


def test_weekly_check_includes_due_sources_when_older_than_seven_days(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    state_path = tmp_path / "state.json"
    _write_companies_yaml(companies_path, [_sample_company(name="Due Co")])
    state_path.write_text(
        json.dumps(
            {
                "sources": {
                    "due co": {
                        "last_health_check_at": "2020-01-01T00:00:00Z"
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Due Co")])
    record_source_observation(
        connection,
        company_name="Due Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    record_source_observation(
        connection,
        company_name="Due Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    fetcher = _fake_fetcher(
        {
            "https://careers.example.com/jobs": (
                "https://careers.example.com/jobs",
                '<a href="/careers">Careers</a>',
                200,
            ),
            "https://careers.example.com": (
                "https://careers.example.com",
                '<a href="/careers">Careers</a>',
                200,
            ),
            "https://careers.example.com/careers": (
                "https://careers.example.com/careers",
                "<html></html>",
                200,
            ),
        }
    )

    candidates = weekly_source_check(
        output_path=tmp_path / "weekly.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        fetch_page=fetcher,
        robots_allowed=_allow_all,
    )

    assert candidates


def test_force_overrides_min_days_between_checks(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    state_path = tmp_path / "state.json"
    _write_companies_yaml(companies_path, [_sample_company(name="Force Co")])
    state_path.write_text(
        json.dumps(
            {
                "sources": {
                    "force co": {
                        "last_health_check_at": "2099-01-01T00:00:00Z"
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Force Co")])
    record_source_observation(
        connection,
        company_name="Force Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    record_source_observation(
        connection,
        company_name="Force Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )

    candidates = weekly_source_check(
        output_path=tmp_path / "weekly.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        force=True,
        fetch_page=_fake_fetcher(
            {
                "https://careers.example.com/jobs": (
                    "https://careers.example.com/jobs",
                    "<html></html>",
                    200,
                )
            }
        ),
        robots_allowed=_allow_all,
    )

    assert candidates


def test_source_health_state_file_is_updated_after_check(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    state_path = tmp_path / "state.json"
    _write_companies_yaml(companies_path, [_sample_company(name="State Co")])
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="State Co")])
    record_source_observation(
        connection,
        company_name="State Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )
    record_source_observation(
        connection,
        company_name="State Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="error",
        error="timeout",
    )

    weekly_source_check(
        output_path=tmp_path / "weekly.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        force=True,
        fetch_page=_fake_fetcher(
            {
                "https://careers.example.com/jobs": (
                    "https://careers.example.com/jobs",
                    "<html></html>",
                    200,
                )
            }
        ),
        robots_allowed=_allow_all,
    )

    payload = json.loads(state_path.read_text(encoding="utf-8"))
    assert "state co" in payload["sources"]
    assert payload["sources"]["state co"]["last_health_check_at"]


def test_refresh_uses_pending_intervention_problem_signal(tmp_path: Path) -> None:
    companies_path = tmp_path / "companies.yaml"
    db_path = tmp_path / "job_discovery.db"
    _write_companies_yaml(companies_path, [_sample_company(name="Intervention Co")])
    connection = initialize_database(db_path)
    upsert_companies(connection, [_sample_company(name="Intervention Co")])
    record_source_observation(
        connection,
        company_name="Intervention Co",
        source_name="company-careers",
        source_mode="browser_allowed",
        careers_url="https://careers.example.com/jobs",
        collector="browser_after_jsonld",
        status="paused",
        intervention_required=True,
    )
    create_intervention(
        connection,
        intervention_type="browser_pause",
        company_name="Intervention Co",
        source_url="https://careers.example.com/jobs",
        reason="cookie_blocked",
    )

    candidates = refresh_sources(
        output_path=tmp_path / "refresh.yaml",
        companies_path=companies_path,
        db_path=db_path,
        state_path=tmp_path / "state.json",
        only_problem_sources=True,
        fetch_page=_fake_fetcher(
            {
                "https://careers.example.com/jobs": (
                    "https://careers.example.com/jobs",
                    "<html></html>",
                    200,
                )
            }
        ),
        robots_allowed=_allow_all,
    )

    assert candidates
