"""Excel importer for company discovery data."""

from __future__ import annotations

import argparse
import re
import sys
from importlib import import_module
from pathlib import Path
from typing import Literal
from urllib.parse import urlparse

import yaml
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

try:
    from classifier.ats_detector import normalize_ats_hint
    from classifier.source_classifier import classify_source
except ModuleNotFoundError:  # pragma: no cover
    # Support running the importer as `python -m src.importer.excel_importer`.
    PROJECT_ROOT = Path(__file__).resolve().parents[2]
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    normalize_ats_hint = import_module("src.classifier.ats_detector").normalize_ats_hint
    classify_source = import_module("src.classifier.source_classifier").classify_source

ALLOWED_CATEGORY = "Bank/Market"
ALLOWED_SECTOR = "IT Consulting & Systems Integrators"
DEFAULT_INPUT_PATH = Path("data/input/Rishi canada companies list (1).xlsx")
DEFAULT_OUTPUT_PATH = Path("config/companies.yaml")
DEFAULT_SHEET_NAME = "Companies"
SOURCE_MODE_API_ALLOWED = "api_allowed"
SOURCE_MODE_BROWSER_ALLOWED = "browser_allowed"
SOURCE_MODE_HUMAN_IN_LOOP = "human_in_loop"
SOURCE_MODE_MANUAL_ONLY = "manual_only"
SOURCE_MODE_NEEDS_URL = "needs_url"
SOURCE_MODE_AVOID = "avoid"
ATS_HINTS = (
    "greenhouse",
    "lever",
    "ashby",
    "smartrecruiters",
    "workday",
    "successfactors",
    "oraclecloud",
    "oracle_hcm",
    "oracle",
    "icims",
    "phenom",
    "ultipro",
)


class CompanyConfig(BaseModel):
    """Normalized company record written to YAML."""

    model_config = ConfigDict(extra="ignore")

    name: str
    sector: str
    category: str
    careers_url: str | None = None
    website_category: str | None = None
    canada_hubs_notes: str | None = None
    role_families: list[str] = Field(default_factory=list)
    keywords: list[str] = Field(default_factory=list)
    priority: str | None = None
    monitoring_hint: str | None = None
    status: str | None = None
    source_mode: Literal[
        "api_allowed",
        "browser_allowed",
        "human_in_loop",
        "manual_only",
        "needs_url",
        "avoid",
    ]
    ats_hint: str | None = None


def clean_text(value: object) -> str | None:
    """Return a trimmed string or None for empty cells."""

    if value is None:
        return None
    text = str(value).strip()
    return text or None


def split_values(value: object, *, pattern: str) -> list[str]:
    """Split a delimited cell into a cleaned list."""

    text = clean_text(value)
    if not text:
        return []
    parts = [part.strip() for part in re.split(pattern, text)]
    return [part for part in parts if part]


def is_real_url(value: object) -> bool:
    """Return True when the value looks like a real http(s) URL."""

    text = clean_text(value)
    if not text:
        return False
    parsed = urlparse(text)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def extract_ats_hint(website_category: str | None) -> str | None:
    """Extract a normalized ATS hint from the website category text."""

    if not website_category:
        return None
    lower_text = website_category.lower()
    for hint in ATS_HINTS:
        if hint in lower_text:
            return normalize_ats_hint(hint)
    return None


def classify_company_source_mode(
    *,
    company_name: str,
    careers_url: str | None,
    website_category: str | None,
    ats_hint: str | None,
) -> str:
    """Return the normalized source mode for one company record."""

    if not careers_url:
        return SOURCE_MODE_NEEDS_URL

    result = classify_source(
        {
            "name": company_name,
            "source_name": website_category or company_name,
            "website_category": website_category,
            "careers_url": careers_url,
            "ats_hint": ats_hint,
        }
    )
    return result.source_mode


def should_include_row(
    category: str | None,
    sector: str | None,
    company: str | None = None,
) -> bool:
    """Return True when the row matches the import filters."""

    if company and company.strip().lower() == "ateko":
        return True
    return (category or "").strip().lower() == ALLOWED_CATEGORY.lower() or (
        sector or ""
    ).strip().lower() == ALLOWED_SECTOR.lower()


def load_company_configs(
    workbook_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[CompanyConfig]:
    """Read the Excel workbook and return normalized company configs."""

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {workbook_path}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(header) or "" for header in next(rows)]

    configs: list[CompanyConfig] = []
    for row in rows:
        row_data = {
            headers[index]: row[index] if index < len(row) else None
            for index in range(len(headers))
        }
        company = clean_text(row_data.get("Company"))
        category = clean_text(row_data.get("Category"))
        sector = clean_text(row_data.get("Sector"))

        if not should_include_row(category, sector, company):
            continue
        if not company:
            continue

        careers_url_raw = clean_text(row_data.get("Careers page URL (fill in)"))
        careers_url = careers_url_raw if is_real_url(careers_url_raw) else None
        website_category = clean_text(row_data.get("website category"))
        ats_hint = extract_ats_hint(website_category)

        configs.append(
            CompanyConfig(
                name=company,
                sector=sector or "",
                category=category or "",
                careers_url=careers_url,
                website_category=website_category,
                canada_hubs_notes=clean_text(row_data.get("Canada hubs / notes")),
                role_families=split_values(row_data.get("Role families"), pattern=r"[\/,;|]+"),
                keywords=split_values(
                    row_data.get("Suggested search keywords"),
                    pattern=r"[\/,;|]+",
                ),
                priority=clean_text(row_data.get("Priority")),
                monitoring_hint=clean_text(row_data.get("Monitoring hint")),
                status=clean_text(row_data.get("Status")),
                source_mode=classify_company_source_mode(
                    company_name=company,
                    careers_url=careers_url,
                    website_category=website_category,
                    ats_hint=ats_hint,
                ),
                ats_hint=ats_hint,
            )
        )

    return configs


def build_companies_payload(configs: list[CompanyConfig]) -> dict[str, list[dict[str, object]]]:
    """Convert company configs into a YAML-friendly payload."""

    return {
        "companies": [config.model_dump(exclude_none=True) for config in configs],
    }


def update_company_record_in_yaml(
    output_path: Path,
    *,
    company_name: str,
    updates: dict[str, object],
) -> dict[str, object]:
    """Update one company record in config/companies.yaml and return the new record."""

    payload = yaml.safe_load(output_path.read_text(encoding="utf-8")) or {}
    companies = payload.get("companies", [])
    if not isinstance(companies, list):
        raise ValueError(f"Expected 'companies' to be a list in {output_path}")

    for company in companies:
        if not isinstance(company, dict):
            continue
        if str(company.get("name") or "").strip() != company_name:
            continue

        for field, value in updates.items():
            if value is None:
                company.pop(field, None)
            else:
                company[field] = value

        output_path.write_text(
            yaml.safe_dump(payload, sort_keys=False, allow_unicode=True, default_flow_style=False),
            encoding="utf-8",
        )
        return company

    raise ValueError(f"Company {company_name!r} not found in {output_path}")


def write_companies_yaml(configs: list[CompanyConfig], output_path: Path) -> None:
    """Write the normalized companies payload to YAML."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = build_companies_payload(configs)
    output_path.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True, default_flow_style=False),
        encoding="utf-8",
    )


def generate_companies_yaml(
    workbook_path: Path = DEFAULT_INPUT_PATH,
    output_path: Path = DEFAULT_OUTPUT_PATH,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[CompanyConfig]:
    """Generate the YAML output file and return the loaded configs."""

    configs = load_company_configs(workbook_path, sheet_name=sheet_name)
    write_companies_yaml(configs, output_path)
    return configs


def build_parser() -> argparse.ArgumentParser:
    """Build the command-line parser."""

    parser = argparse.ArgumentParser(
        description="Generate config/companies.yaml from the Excel workbook.",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT_PATH,
        help="Input Excel workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help="Output YAML path.",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help="Worksheet name to read.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""

    args = build_parser().parse_args(argv)
    generate_companies_yaml(args.input, args.output, sheet_name=args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
