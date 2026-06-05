"""Review-first company/source onboarding and refresh helpers."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
from collections import Counter
from datetime import UTC, datetime
from importlib import import_module
from pathlib import Path
from typing import Any, Literal
from urllib.parse import urlparse

import yaml
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

try:
    from classifier.ats_detector import API_ALLOWED_ATS, detect_ats_type
    from classifier.source_classifier import classify_source
    from importer.apply_career_urls import load_yaml_companies
    from importer.excel_importer import (
        clean_text,
        is_real_url,
        load_company_configs,
        update_company_record_in_yaml,
    )
    from onboarding.live_discovery import discover_live_candidates
    from storage.db import get_intervention_queue, get_source_status_rows, initialize_database
except ModuleNotFoundError:  # pragma: no cover
    ats_module = import_module("src.classifier.ats_detector")
    source_classifier_module = import_module("src.classifier.source_classifier")
    apply_urls_module = import_module("src.importer.apply_career_urls")
    excel_importer_module = import_module("src.importer.excel_importer")
    live_discovery_module = import_module("src.onboarding.live_discovery")
    storage_module = import_module("src.storage.db")
    API_ALLOWED_ATS = ats_module.API_ALLOWED_ATS
    detect_ats_type = ats_module.detect_ats_type
    classify_source = source_classifier_module.classify_source
    load_yaml_companies = apply_urls_module.load_yaml_companies
    clean_text = excel_importer_module.clean_text
    is_real_url = excel_importer_module.is_real_url
    load_company_configs = excel_importer_module.load_company_configs
    update_company_record_in_yaml = excel_importer_module.update_company_record_in_yaml
    discover_live_candidates = live_discovery_module.discover_live_candidates
    get_intervention_queue = storage_module.get_intervention_queue
    get_source_status_rows = storage_module.get_source_status_rows
    initialize_database = storage_module.initialize_database

DEFAULT_INPUT_PATH = PROJECT_ROOT / "data" / "input" / "companies.txt"
DEFAULT_LARGE_LIST_INPUT_PATH = (
    PROJECT_ROOT / "data" / "input" / "Rishi canada companies list (1).xlsx"
)
DEFAULT_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "source-onboarding-candidates.yaml"
)
DEFAULT_READINESS_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "company-input-readiness.csv"
)
DEFAULT_LARGE_LIST_CANDIDATES_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "large-list-source-candidates.yaml"
)
DEFAULT_NEEDS_WEBSITE_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "large-list-needs-website-input.csv"
)
DEFAULT_LARGE_LIST_REPORT_PATH = (
    PROJECT_ROOT / "docs" / "large-company-list-readiness-report.md"
)
DEFAULT_REFRESH_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "source-refresh-candidates.yaml"
)
DEFAULT_WEEKLY_OUTPUT_PATH = (
    PROJECT_ROOT / "data" / "exports" / "weekly-source-refresh-candidates.yaml"
)
DEFAULT_HEALTH_STATE_PATH = PROJECT_ROOT / "data" / "exports" / "source-health-state.json"
DEFAULT_COMPANIES_PATH = PROJECT_ROOT / "config" / "companies.yaml"
DEFAULT_STARTER_PATH = PROJECT_ROOT / "config" / "starter_career_urls.yaml"
DEFAULT_DATABASE_PATH = PROJECT_ROOT / "data" / "job_discovery.db"
DEFAULT_REFERENCE_WORKBOOKS = (
    PROJECT_ROOT / "data" / "input" / "Rishi canada companies list (1).xlsx",
    PROJECT_ROOT / "data" / "input" / "rishi" / "companies.xlsx",
)
RESTRICTED_SOURCE_MODES = {"manual_only"}
CSV_NAME_COLUMNS = ("company_name", "company", "name")
CSV_WEBSITE_COLUMNS = ("website_url", "official_website", "company_website", "website")
CSV_CAREERS_COLUMNS = ("careers_url", "career_url", "job_board_url")
REVIEWABLE_CONFIDENCE = {"medium", "low"}
HARD_ERROR_MARKERS = (
    "timeout",
    "err_name_not_resolved",
    "dns",
    "404",
    "410",
    "not found",
    "name_not_resolved",
)


class CompanyInput(BaseModel):
    """One onboarding input row."""

    model_config = ConfigDict(extra="ignore")

    company_name: str
    website_url: str | None = None
    careers_url: str | None = None


class OnboardingCandidate(BaseModel):
    """Reviewable candidate source entry."""

    model_config = ConfigDict(extra="ignore")

    company_name: str
    candidate_official_website: str | None = None
    candidate_careers_url: str | None = None
    candidate_job_board_url: str | None = None
    detected_ats_type: str | None = None
    suggested_source_mode: str
    confidence: Literal["high", "medium", "low"]
    needs_review: bool
    reason: str
    evidence: list[str] = Field(default_factory=list)
    approved: bool = False
    sector: str | None = None
    category: str | None = None
    website_category: str | None = None
    ats_hint: str | None = None
    canada_hubs_notes: str | None = None
    role_families: list[str] = Field(default_factory=list)
    keywords: list[str] = Field(default_factory=list)
    priority: str | None = None
    monitoring_hint: str | None = None
    status: str | None = None
    current_careers_url: str | None = None
    current_source_mode: str | None = None
    current_ats_type: str | None = None
    current_status_or_last_error: str | None = None
    suggested_action: str | None = None


class SpreadsheetCompanyRecord(BaseModel):
    """One large-list spreadsheet row with hyperlink-aware fields."""

    model_config = ConfigDict(extra="ignore")

    company_name: str
    spreadsheet_career_display_text: str | None = None
    spreadsheet_career_url_or_hyperlink: str | None = None
    website_category: str | None = None
    sector: str | None = None
    category: str | None = None
    canada_hubs_notes: str | None = None
    role_families: list[str] = Field(default_factory=list)
    keywords: list[str] = Field(default_factory=list)
    early_career_pipeline: str | None = None
    priority: str | None = None
    monitoring_hint: str | None = None
    status: str | None = None
    notes: str | None = None


class ReadinessAuditRow(BaseModel):
    """Spreadsheet readiness audit row."""

    model_config = ConfigDict(extra="ignore")

    company_name: str
    sector: str | None = None
    category: str | None = None
    priority: str | None = None
    status: str | None = None
    existing_config_match: bool = False
    existing_config_url: str | None = None
    spreadsheet_career_display_text: str | None = None
    spreadsheet_career_url_or_hyperlink: str | None = None
    starter_url_match: bool = False
    starter_url: str | None = None
    usable_url_available: bool = False
    detected_ats_type: str | None = None
    suggested_source_mode: str | None = None
    readiness_status: str
    recommended_next_action: str
    notes: str | None = None
    existing_source_mode: str | None = None
    existing_ats_hint: str | None = None
    existing_status: str | None = None


def _normalize_company_name(value: str | None) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _normalize_url(url: str | None) -> str | None:
    if not is_real_url(url):
        return None
    parsed = urlparse(str(url))
    path = parsed.path or "/"
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{parsed.scheme.lower()}://{parsed.netloc.lower()}{path}{query}"


def _root_url(url: str | None) -> str | None:
    normalized = _normalize_url(url)
    if not normalized:
        return None
    parsed = urlparse(normalized)
    return f"{parsed.scheme}://{parsed.netloc}"


def _candidate_sort_key(candidate: OnboardingCandidate) -> tuple[str, str, str]:
    url = candidate.candidate_job_board_url or candidate.candidate_careers_url or ""
    return (_normalize_company_name(candidate.company_name), candidate.reason, url)


def _candidate_job_board_url(url: str | None, ats_type: str | None) -> str | None:
    normalized = _normalize_url(url)
    if not normalized:
        return None
    if ats_type:
        return normalized
    hostname = urlparse(normalized).netloc.lower()
    if hostname.startswith(("jobs.", "careers.")):
        return normalized
    return None


def _candidate_official_website(url: str | None, ats_type: str | None) -> str | None:
    normalized = _normalize_url(url)
    if not normalized:
        return None
    if ats_type in API_ALLOWED_ATS or ats_type == "restricted_board":
        return None
    return _root_url(normalized)


def _safe_existing_list(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    return []


def _load_reference_companies(
    workbook_paths: tuple[Path, ...] = DEFAULT_REFERENCE_WORKBOOKS,
) -> dict[str, dict[str, Any]]:
    reference: dict[str, dict[str, Any]] = {}
    for workbook_path in workbook_paths:
        if not workbook_path.exists():
            continue
        for config in load_company_configs(workbook_path):
            key = _normalize_company_name(config.name)
            reference.setdefault(key, config.model_dump(exclude_none=True))
    return reference


def _load_starter_index(starter_path: Path) -> dict[str, dict[str, Any]]:
    starter_index: dict[str, dict[str, Any]] = {}
    if not starter_path.exists():
        return starter_index
    for company in load_yaml_companies(starter_path):
        if not isinstance(company, dict):
            continue
        key = _normalize_company_name(company.get("name"))
        if key:
            starter_index[key] = company
    return starter_index


def _load_config_index(companies_path: Path) -> dict[str, dict[str, Any]]:
    config_index: dict[str, dict[str, Any]] = {}
    if not companies_path.exists():
        return config_index
    for company in load_yaml_companies(companies_path):
        if not isinstance(company, dict):
            continue
        key = _normalize_company_name(company.get("name"))
        if key:
            config_index[key] = company
    return config_index


def _classify_candidate_source(
    *,
    company_name: str,
    careers_url: str | None,
    website_category: str | None,
    ats_hint: str | None,
    current_source_mode: str | None = None,
) -> tuple[str | None, str]:
    ats_type = detect_ats_type(
        careers_url,
        ats_hint=ats_hint,
        website_category=website_category,
    )
    classification = classify_source(
        {
            "name": company_name,
            "source_name": website_category or company_name,
            "website_category": website_category,
            "careers_url": careers_url,
            "ats_hint": ats_hint,
            "source_mode": current_source_mode,
        }
    )
    return ats_type, classification.source_mode


def _build_candidate(
    *,
    company_name: str,
    source_record: dict[str, Any] | None,
    careers_url: str | None,
    official_website: str | None = None,
    confidence: str,
    reason: str,
    evidence: list[str],
    current_source_mode: str | None = None,
    current_careers_url: str | None = None,
    current_ats_type: str | None = None,
    current_status_or_last_error: str | None = None,
    suggested_action: str | None = None,
) -> OnboardingCandidate:
    source_record = source_record or {}
    normalized_url = _normalize_url(careers_url)
    normalized_website = _normalize_url(official_website)
    website_category = clean_text(source_record.get("website_category"))
    ats_hint = clean_text(source_record.get("ats_hint"))
    ats_type, source_mode = _classify_candidate_source(
        company_name=company_name,
        careers_url=normalized_url,
        website_category=website_category,
        ats_hint=ats_hint,
        current_source_mode=current_source_mode,
    )

    resolved_confidence: Literal["high", "medium", "low"]
    if not normalized_url:
        resolved_confidence = "low"
    elif ats_type in API_ALLOWED_ATS:
        resolved_confidence = "high"
    elif confidence == "high":
        resolved_confidence = "high"
    elif confidence == "medium":
        resolved_confidence = "medium"
    else:
        resolved_confidence = "low"

    if source_mode == "manual_only":
        resolved_confidence = "low"
        needs_review = True
        if reason != "restricted_board_candidate":
            reason = "restricted_board_candidate"
        if "restricted board detected" not in evidence:
            evidence.append("restricted board detected")
    else:
        needs_review = resolved_confidence in REVIEWABLE_CONFIDENCE or not normalized_url

    return OnboardingCandidate(
        company_name=company_name,
        candidate_official_website=normalized_website
        or _candidate_official_website(normalized_url, ats_type),
        candidate_careers_url=normalized_url,
        candidate_job_board_url=_candidate_job_board_url(normalized_url, ats_type),
        detected_ats_type=ats_type,
        suggested_source_mode=source_mode,
        confidence=resolved_confidence,
        needs_review=needs_review,
        reason=reason,
        evidence=evidence,
        approved=False,
        sector=clean_text(source_record.get("sector")),
        category=clean_text(source_record.get("category")),
        website_category=website_category,
        ats_hint=ats_hint,
        canada_hubs_notes=clean_text(source_record.get("canada_hubs_notes")),
        role_families=_safe_existing_list(source_record.get("role_families")),
        keywords=_safe_existing_list(source_record.get("keywords")),
        priority=clean_text(source_record.get("priority")),
        monitoring_hint=clean_text(source_record.get("monitoring_hint")),
        status=clean_text(source_record.get("status")),
        current_careers_url=_normalize_url(current_careers_url),
        current_source_mode=current_source_mode,
        current_ats_type=current_ats_type,
        current_status_or_last_error=current_status_or_last_error,
        suggested_action=suggested_action,
    )


def _candidate_from_existing_config(
    company_name: str,
    company: dict[str, Any],
) -> OnboardingCandidate:
    return _build_candidate(
        company_name=company_name,
        source_record=company,
        careers_url=company.get("careers_url"),
        confidence="high",
        reason="existing_config_match",
        evidence=[
            "source=existing_config",
            f"existing_source_mode={company.get('source_mode') or 'unknown'}",
        ],
        current_source_mode=company.get("source_mode"),
        suggested_action="keep_existing",
    )


def _candidate_from_starter(
    company_name: str,
    starter: dict[str, Any],
    reference_company: dict[str, Any] | None,
) -> OnboardingCandidate:
    starter_confidence = str(starter.get("confidence") or "medium").strip().lower()
    confidence = starter_confidence if starter_confidence in {"high", "medium"} else "low"
    evidence = ["source=starter_file"]
    notes = clean_text(starter.get("notes"))
    if notes:
        evidence.append(f"starter_notes={notes}")
    return _build_candidate(
        company_name=company_name,
        source_record=reference_company or {},
        careers_url=starter.get("careers_url"),
        confidence=confidence,
        reason="starter_career_url_match",
        evidence=evidence,
        current_source_mode=(reference_company or {}).get("source_mode"),
    )


def _candidate_from_reference_import(
    company_name: str,
    reference_company: dict[str, Any],
) -> OnboardingCandidate:
    return _build_candidate(
        company_name=company_name,
        source_record=reference_company,
        careers_url=reference_company.get("careers_url"),
        confidence="medium",
        reason="reference_import_match",
        evidence=["source=workbook"],
        current_source_mode=reference_company.get("source_mode"),
    )


def _candidate_from_provided_careers_url(
    company_input: CompanyInput,
) -> OnboardingCandidate:
    evidence = ["source=provided_careers_url"]
    if company_input.careers_url:
        evidence.append(f"link_url={company_input.careers_url}")
    return _build_candidate(
        company_name=company_input.company_name,
        source_record={},
        careers_url=company_input.careers_url,
        official_website=company_input.website_url,
        confidence="medium",
        reason="provided_careers_url",
        evidence=evidence,
    )


def _missing_candidate(
    company_name: str,
    *,
    official_website: str | None = None,
    reason: str = "missing_candidate_url",
    evidence: list[str] | None = None,
) -> OnboardingCandidate:
    return OnboardingCandidate(
        company_name=company_name,
        candidate_official_website=_normalize_url(official_website),
        candidate_careers_url=None,
        candidate_job_board_url=None,
        detected_ats_type=None,
        suggested_source_mode="needs_url",
        confidence="low",
        needs_review=True,
        reason=reason,
        evidence=evidence
        or [
            "source=internal_lookup",
            "no matching source found in config/companies.yaml",
            "no matching starter career URL found",
            "no matching reference import source found",
        ],
        approved=False,
        suggested_action="manual_review_required",
    )


def _candidate_from_live_discovery(
    *,
    company_input: CompanyInput,
    finding: Any,
    source_record: dict[str, Any] | None,
    current_careers_url: str | None = None,
    current_source_mode: str | None = None,
    current_ats_type: str | None = None,
    current_status_or_last_error: str | None = None,
) -> OnboardingCandidate:
    candidate_url = finding.url
    official_website = company_input.website_url or _root_url(candidate_url or current_careers_url)
    suggested_action = "manual_review_required"
    if current_careers_url and _normalize_url(candidate_url) == _normalize_url(current_careers_url):
        suggested_action = "keep_existing"
    elif candidate_url and _normalize_url(candidate_url) != _normalize_url(current_careers_url):
        suggested_action = "replace_with_candidate"
    if finding.restricted:
        suggested_action = "manual_review_required"
    if not candidate_url:
        suggested_action = "mark_needs_url"
    return _build_candidate(
        company_name=company_input.company_name,
        source_record=source_record,
        careers_url=candidate_url,
        official_website=official_website,
        confidence=finding.confidence,
        reason=finding.reason,
        evidence=finding.evidence,
        current_source_mode=current_source_mode,
        current_careers_url=current_careers_url,
        current_ats_type=current_ats_type,
        current_status_or_last_error=current_status_or_last_error,
        suggested_action=suggested_action,
    )


def load_company_inputs(input_path: Path) -> list[CompanyInput]:
    """Load structured onboarding rows from TXT, CSV, or XLSX."""

    suffix = input_path.suffix.lower()
    if suffix == ".txt":
        items: list[CompanyInput] = []
        for line in input_path.read_text(encoding="utf-8").splitlines():
            text = line.strip()
            if not text:
                continue
            if "|" in text:
                parts = [part.strip() for part in text.split("|")]
                company_name = parts[0]
                website_url = parts[1] if len(parts) > 1 else None
                careers_url = parts[2] if len(parts) > 2 else None
                items.append(
                    CompanyInput(
                        company_name=company_name,
                        website_url=website_url or None,
                        careers_url=careers_url or None,
                    )
                )
                continue
            items.append(CompanyInput(company_name=text))
        return items

    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            if not reader.fieldnames:
                csv_file.seek(0)
                plain_reader = csv.reader(csv_file)
                return [
                    CompanyInput(company_name=row[0].strip())
                    for row in plain_reader
                    if row and row[0].strip()
                ]
            name_field = next(
                (field for field in CSV_NAME_COLUMNS if field in reader.fieldnames),
                None,
            )
            website_field = next(
                (field for field in CSV_WEBSITE_COLUMNS if field in reader.fieldnames),
                None,
            )
            careers_field = next(
                (field for field in CSV_CAREERS_COLUMNS if field in reader.fieldnames),
                None,
            )
            if name_field is None:
                raise ValueError("CSV input must include a company name column.")
            return [
                CompanyInput(
                    company_name=str(row.get(name_field) or "").strip(),
                    website_url=str(row.get(website_field) or "").strip() or None
                    if website_field
                    else None,
                    careers_url=str(row.get(careers_field) or "").strip() or None
                    if careers_field
                    else None,
                )
                for row in reader
                if str(row.get(name_field) or "").strip()
            ]

    if suffix == ".xlsx":
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean_text(value) or "" for value in rows[0]]
        name_index = next(
            (headers.index(field) for field in ("Company", "company_name", "company", "name")
             if field in headers),
            None,
        )
        if name_index is None:
            raise ValueError("XLSX input must include a company name column.")
        website_index = next(
            (headers.index(field) for field in CSV_WEBSITE_COLUMNS if field in headers),
            None,
        )
        careers_index = next(
            (headers.index(field) for field in CSV_CAREERS_COLUMNS if field in headers),
            None,
        )
        items: list[CompanyInput] = []
        for row in rows[1:]:
            company_name = str(row[name_index] or "").strip()
            if not company_name:
                continue
            website_url = None
            careers_url = None
            if website_index is not None and website_index < len(row):
                website_url = str(row[website_index] or "").strip() or None
            if careers_index is not None and careers_index < len(row):
                careers_url = str(row[careers_index] or "").strip() or None
            items.append(
                CompanyInput(
                    company_name=company_name,
                    website_url=website_url,
                    careers_url=careers_url,
                )
            )
        return items

    raise ValueError(f"Unsupported onboarding input type: {input_path.suffix}")


def load_company_names(input_path: Path) -> list[str]:
    """Backward-compatible company-name loader."""

    return [item.company_name for item in load_company_inputs(input_path)]


def _normalize_match_key(value: str | None) -> str:
    text = str(value or "").strip().lower().replace("&", " and ")
    cleaned = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(cleaned.split())


def _company_match_keys(value: str | None) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    keys: list[str] = []

    def add(text: str) -> None:
        normalized = _normalize_match_key(text)
        if normalized and normalized not in keys:
            keys.append(normalized)

    add(raw)
    collapsed = re.sub(r"\([^)]*\)", "", raw).strip()
    if collapsed and collapsed != raw:
        add(collapsed)
    for alias in re.findall(r"\(([^)]*)\)", raw):
        add(alias)
    for part in re.split(r"/", raw):
        stripped = part.strip()
        if stripped and stripped != raw:
            add(stripped)
    return keys


def _build_match_index(companies: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = {}
    for company in companies:
        if not isinstance(company, dict):
            continue
        for key in _company_match_keys(company.get("name")):
            index.setdefault(key, []).append(company)
    return index


def _resolve_index_match(
    index: dict[str, list[dict[str, Any]]],
    company_name: str,
) -> tuple[dict[str, Any] | None, bool]:
    exact_key = _normalize_match_key(company_name)
    exact_matches = {
        str(item.get("name") or ""): item
        for item in index.get(exact_key, [])
    }
    if len(exact_matches) == 1:
        return next(iter(exact_matches.values())), False
    if len(exact_matches) > 1:
        return None, True

    matches: dict[str, dict[str, Any]] = {}
    for key in _company_match_keys(company_name):
        for item in index.get(key, []):
            matches[str(item.get("name") or "")] = item
    if len(matches) == 1:
        return next(iter(matches.values())), False
    if len(matches) > 1:
        return None, True
    return None, False


def _parse_multivalue_text(value: str | None) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[\n,;|]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def _read_hyperlink_target(cell: Any) -> str | None:
    if cell is None or cell.hyperlink is None:
        return None
    target = clean_text(getattr(cell.hyperlink, "target", None))
    return target or None


def load_spreadsheet_company_records(input_path: Path) -> list[SpreadsheetCompanyRecord]:
    """Load the large-list spreadsheet with hyperlink-aware careers URL handling."""

    workbook = load_workbook(input_path, read_only=False, data_only=False)
    if "Companies" in workbook.sheetnames:
        sheet = workbook["Companies"]
    else:
        sheet = workbook[workbook.sheetnames[0]]
    headers = [clean_text(cell.value) or "" for cell in sheet[1]]
    header_lookup = {header: index for index, header in enumerate(headers)}
    required_header = "Company"
    if required_header not in header_lookup:
        raise ValueError("Spreadsheet must include a Company column.")

    def index_for(*candidates: str) -> int | None:
        return next((header_lookup[item] for item in candidates if item in header_lookup), None)

    company_index = header_lookup[required_header]
    careers_index = index_for("Careers page URL (fill in)")
    website_category_index = index_for("website category")
    sector_index = index_for("Sector")
    category_index = index_for("Category")
    hubs_index = index_for("Canada hubs / notes")
    role_families_index = index_for("Role families")
    keywords_index = index_for("Suggested search keywords")
    early_pipeline_index = index_for("Early-career pipeline")
    priority_index = index_for("Priority")
    monitoring_index = index_for("Monitoring hint")
    status_index = index_for("Status")
    notes_index = index_for("Notes")

    records: list[SpreadsheetCompanyRecord] = []
    for row in sheet.iter_rows(min_row=2):
        company_name = clean_text(row[company_index].value)
        if not company_name:
            continue

        careers_cell = row[careers_index] if careers_index is not None else None
        display_text = clean_text(careers_cell.value if careers_cell is not None else None)
        hyperlink_target = _read_hyperlink_target(careers_cell)
        spreadsheet_url = hyperlink_target if is_real_url(hyperlink_target) else None

        records.append(
            SpreadsheetCompanyRecord(
                company_name=company_name,
                spreadsheet_career_display_text=display_text,
                spreadsheet_career_url_or_hyperlink=spreadsheet_url,
                website_category=clean_text(
                    row[website_category_index].value
                    if website_category_index is not None
                    else None
                ),
                sector=clean_text(
                    row[sector_index].value if sector_index is not None else None
                ),
                category=clean_text(
                    row[category_index].value if category_index is not None else None
                ),
                canada_hubs_notes=clean_text(
                    row[hubs_index].value if hubs_index is not None else None
                ),
                role_families=_parse_multivalue_text(
                    row[role_families_index].value if role_families_index is not None else None
                ),
                keywords=_parse_multivalue_text(
                    row[keywords_index].value if keywords_index is not None else None
                ),
                early_career_pipeline=clean_text(
                    row[early_pipeline_index].value if early_pipeline_index is not None else None
                ),
                priority=clean_text(
                    row[priority_index].value if priority_index is not None else None
                ),
                monitoring_hint=clean_text(
                    row[monitoring_index].value if monitoring_index is not None else None
                ),
                status=clean_text(row[status_index].value if status_index is not None else None),
                notes=clean_text(row[notes_index].value if notes_index is not None else None),
            )
        )
    return records


def _spreadsheet_record_to_source_record(record: SpreadsheetCompanyRecord) -> dict[str, Any]:
    return {
        "name": record.company_name,
        "sector": record.sector,
        "category": record.category,
        "website_category": record.website_category,
        "canada_hubs_notes": record.canada_hubs_notes,
        "role_families": record.role_families,
        "keywords": record.keywords,
        "priority": record.priority,
        "monitoring_hint": record.monitoring_hint,
        "status": record.status,
    }


def _best_candidate_url(
    *,
    existing_company: dict[str, Any] | None,
    spreadsheet_record: SpreadsheetCompanyRecord,
    starter_company: dict[str, Any] | None,
) -> tuple[str | None, str]:
    existing_url = clean_text((existing_company or {}).get("careers_url"))
    if is_real_url(existing_url):
        return str(existing_url), "existing_config"
    spreadsheet_url = spreadsheet_record.spreadsheet_career_url_or_hyperlink
    if is_real_url(spreadsheet_url):
        return str(spreadsheet_url), "spreadsheet_hyperlink"
    starter_url = clean_text((starter_company or {}).get("careers_url"))
    if is_real_url(starter_url):
        return str(starter_url), "starter_url"
    return None, "none"


def _classify_readiness_status(
    *,
    existing_match: bool,
    ambiguous_match: bool,
    best_url_source: str,
    source_mode: str | None,
    display_text: str | None,
) -> tuple[str, str]:
    if ambiguous_match:
        return "duplicate_or_alias_review", "merge_duplicate"
    if existing_match:
        return "already_configured", "no_action"
    if source_mode == "manual_only":
        return "restricted_manual_only", "skip_or_manual_tracking"
    if best_url_source == "spreadsheet_hyperlink":
        return "ready_with_spreadsheet_url", "review_and_apply"
    if best_url_source == "starter_url":
        return "ready_with_starter_url", "review_and_apply"
    if display_text:
        return "needs_manual_career_url", "manually_find_career_url"
    return "needs_website_url", "add_website_for_live_discovery"


def _write_readiness_csv(rows: list[ReadinessAuditRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "company_name",
        "sector",
        "category",
        "priority",
        "status",
        "existing_config_match",
        "existing_config_url",
        "spreadsheet_career_display_text",
        "spreadsheet_career_url_or_hyperlink",
        "starter_url_match",
        "starter_url",
        "usable_url_available",
        "detected_ats_type",
        "suggested_source_mode",
        "readiness_status",
        "recommended_next_action",
        "notes",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.model_dump(include=set(fieldnames)))


def _write_needs_website_csv(rows: list[ReadinessAuditRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["company_name", "website_url", "notes"])
        writer.writeheader()
        for row in rows:
            if row.readiness_status != "needs_website_url":
                continue
            writer.writerow(
                {
                    "company_name": row.company_name,
                    "website_url": "",
                    "notes": row.notes or "Add company website for later live discovery.",
                }
            )


def _priority_is_high(priority: str | None) -> bool:
    return str(priority or "").strip().lower() == "high"


def _unique_company_names(rows: list[ReadinessAuditRow]) -> list[str]:
    names: list[str] = []
    for row in rows:
        if row.company_name not in names:
            names.append(row.company_name)
    return names


def _format_company_list(names: list[str], limit: int = 20) -> str:
    if not names:
        return "- None"
    visible = names[:limit]
    lines = [f"- {name}" for name in visible]
    remaining = len(names) - len(visible)
    if remaining > 0:
        lines.append(f"- ... and {remaining} more")
    return "\n".join(lines)


def _build_large_list_report(
    *,
    inspected_inputs: tuple[Path, ...],
    rows: list[ReadinessAuditRow],
    candidates: list[OnboardingCandidate],
    output_path: Path,
) -> dict[str, Any]:
    total_companies = len(rows)
    already_configured_count = sum(1 for row in rows if row.existing_config_match)
    usable_url_count = sum(1 for row in rows if row.usable_url_available)
    missing_url_count = sum(1 for row in rows if not row.usable_url_available)
    spreadsheet_hyperlink_count = sum(
        1 for row in rows if row.spreadsheet_career_url_or_hyperlink
    )
    starter_url_match_count = sum(1 for row in rows if row.starter_url_match)
    source_mode_distribution = Counter(
        row.suggested_source_mode or "unknown" for row in rows
    )
    ats_type_distribution = Counter(row.detected_ats_type or "none" for row in rows)

    high_priority_ready_now = [
        row
        for row in rows
        if _priority_is_high(row.priority)
        and row.readiness_status
        in {"already_configured", "ready_with_spreadsheet_url", "ready_with_starter_url"}
    ]
    high_priority_needs_review = [
        row
        for row in rows
        if _priority_is_high(row.priority)
        and row.readiness_status
        in {
            "needs_website_url",
            "needs_manual_career_url",
            "restricted_manual_only",
            "duplicate_or_alias_review",
        }
    ]
    safe_next_batch = [
        row
        for row in rows
        if not row.existing_config_match
        and row.usable_url_available
        and row.suggested_source_mode in {"api_allowed", "browser_allowed", "human_in_loop"}
    ]
    not_ready = [
        row
        for row in rows
        if row.readiness_status
        in {
            "needs_website_url",
            "needs_manual_career_url",
            "restricted_manual_only",
            "duplicate_or_alias_review",
        }
    ]

    report = "\n".join(
        [
            "# Large Company List Readiness Report",
            "",
            "## Verdict",
            "",
            "The 150-company spreadsheet has been audited for MVP input readiness.",
            "Configured companies are separated from additional reviewable source candidates,",
            "and missing/manual-only rows are clearly marked without changing config.",
            "",
            "## Input Files Inspected",
            "",
            *[f"- `{path}`" for path in inspected_inputs],
            "",
            "## Total Companies In Spreadsheet",
            "",
            f"- {total_companies}",
            "",
            "## Already Configured Count",
            "",
            f"- {already_configured_count}",
            "",
            "## Usable URL Count",
            "",
            f"- {usable_url_count}",
            "",
            "## Missing URL Count",
            "",
            f"- {missing_url_count}",
            "",
            "## Spreadsheet Hyperlink Count",
            "",
            f"- {spreadsheet_hyperlink_count}",
            "",
            "## Starter URL Match Count",
            "",
            f"- {starter_url_match_count}",
            "",
            "## Source Mode Distribution",
            "",
            *[f"- `{mode}`: {count}" for mode, count in source_mode_distribution.most_common()],
            "",
            "## ATS Type Distribution",
            "",
            *[
                f"- `{ats_type}`: {count}"
                for ats_type, count in ats_type_distribution.most_common()
            ],
            "",
            "## High-Priority Companies Ready Now",
            "",
            _format_company_list(_unique_company_names(high_priority_ready_now)),
            "",
            "## High-Priority Companies Needing URL/Manual Review",
            "",
            _format_company_list(_unique_company_names(high_priority_needs_review)),
            "",
            "## Companies Safe To Test In Next Batch",
            "",
            _format_company_list(_unique_company_names(safe_next_batch)),
            "",
            "## Companies Not Ready For Testing",
            "",
            _format_company_list(_unique_company_names(not_ready)),
            "",
            "## Recommended Batch Plan",
            "",
            (
                f"- Batch 1: current configured 34 "
                f"(`{already_configured_count}` rows matched config in this audit)"
            ),
            (
                f"- Batch 2: additional companies with confirmed URLs "
                f"(`{len(_unique_company_names(safe_next_batch))}` "
                f"candidates ready for review/apply)"
            ),
            (
                f"- Batch 3: companies needing website URL/live discovery/manual review "
                f"(`{len(_unique_company_names(not_ready))}` rows)"
            ),
            "",
            "## Risks And Limitations",
            "",
            (
                "- Spreadsheet display text without a hyperlink is treated as "
                "unverified and not auto-used as a URL."
            ),
            "- Restricted boards such as LinkedIn, Indeed, and Glassdoor remain manual-only.",
            "- No candidates were auto-applied to `config/companies.yaml`.",
            "- This is a readiness audit only, not a full 150-company discovery run.",
            "- Some alias or duplicate matches may still require human review.",
            "",
            "## Candidate Generation Summary",
            "",
            f"- Reviewable candidates generated: {len(candidates)}",
            (
                f"- High-confidence candidates: "
                f"{sum(1 for item in candidates if item.confidence == 'high')}"
            ),
            (
                f"- Manual-only candidates: "
                f"{sum(1 for item in candidates if item.suggested_source_mode == 'manual_only')}"
            ),
        ]
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(report + "\n", encoding="utf-8")
    return {
        "total_companies": total_companies,
        "already_configured_count": already_configured_count,
        "usable_url_count": usable_url_count,
        "missing_url_count": missing_url_count,
        "spreadsheet_hyperlink_count": spreadsheet_hyperlink_count,
        "starter_url_match_count": starter_url_match_count,
        "source_mode_distribution": dict(source_mode_distribution),
        "ats_type_distribution": dict(ats_type_distribution),
        "candidate_count": len(candidates),
    }


def _dedupe_candidates(
    candidates: list[OnboardingCandidate],
) -> list[OnboardingCandidate]:
    seen: set[tuple[str, str | None, str | None, str]] = set()
    deduped: list[OnboardingCandidate] = []
    for candidate in sorted(candidates, key=_candidate_sort_key):
        key = (
            _normalize_company_name(candidate.company_name),
            candidate.candidate_careers_url,
            candidate.candidate_job_board_url,
            candidate.reason,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(candidate)
    return deduped


def generate_candidates(
    company_names: list[str],
    *,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    starter_path: Path = DEFAULT_STARTER_PATH,
    reference_workbooks: tuple[Path, ...] = DEFAULT_REFERENCE_WORKBOOKS,
) -> list[OnboardingCandidate]:
    """Preserve Task 10 internal-only onboarding for name-only inputs."""

    return generate_candidates_from_records(
        [CompanyInput(company_name=name) for name in company_names],
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=reference_workbooks,
        live_discovery=False,
    )


def generate_candidates_from_records(
    company_inputs: list[CompanyInput],
    *,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    starter_path: Path = DEFAULT_STARTER_PATH,
    reference_workbooks: tuple[Path, ...] = DEFAULT_REFERENCE_WORKBOOKS,
    live_discovery: bool = False,
    max_pages_per_company: int = 8,
    fetch_page: Any = None,
    robots_allowed: Any = None,
) -> list[OnboardingCandidate]:
    """Generate onboarding candidates from structured inputs."""

    config_index = _load_config_index(companies_path)
    starter_index = _load_starter_index(starter_path)
    reference_index = _load_reference_companies(reference_workbooks)
    generated: list[OnboardingCandidate] = []

    for company_input in company_inputs:
        company_name = company_input.company_name.strip()
        if not company_name:
            continue
        key = _normalize_company_name(company_name)
        config_company = config_index.get(key)
        starter_company = starter_index.get(key)
        reference_company = reference_index.get(key)

        if company_input.careers_url and is_real_url(company_input.careers_url):
            generated.append(_candidate_from_provided_careers_url(company_input))
        elif config_company is not None:
            generated.append(_candidate_from_existing_config(company_name, config_company))
        elif starter_company is not None:
            generated.append(
                _candidate_from_starter(
                    company_name,
                    starter_company,
                    reference_company,
                )
            )
        elif reference_company is not None:
            generated.append(_candidate_from_reference_import(company_name, reference_company))
        else:
            generated.append(_missing_candidate(company_name))

        if not live_discovery:
            continue

        start_urls: list[str] = []
        if is_real_url(company_input.website_url):
            start_urls.append(str(company_input.website_url))
        if is_real_url(company_input.careers_url):
            start_urls.append(str(company_input.careers_url))
        if not start_urls and config_company and is_real_url(config_company.get("careers_url")):
            start_urls.append(str(config_company.get("careers_url")))

        if not start_urls:
            if (
                config_company is None
                and starter_company is None
                and reference_company is None
            ):
                generated.append(
                    _missing_candidate(
                        company_name,
                        official_website=company_input.website_url,
                        reason="missing_candidate_url",
                        evidence=[
                            "source=internal_lookup",
                            "company-name-only live discovery requires a provided website URL",
                            "no search provider is configured in this task",
                        ],
                    )
                )
            continue

        source_record = config_company or reference_company or {}
        live_findings = discover_live_candidates(
            company_name=company_name,
            start_urls=start_urls,
            max_pages=max_pages_per_company,
            fetch_page=fetch_page,
            robots_allowed=robots_allowed,
        )
        for finding in live_findings:
            generated.append(
                _candidate_from_live_discovery(
                    company_input=company_input,
                    finding=finding,
                    source_record=source_record,
                )
            )

    return _dedupe_candidates(generated)


def _write_candidates_yaml(candidates: list[OnboardingCandidate], output_path: Path) -> None:
    payload = {
        "candidates": [
            candidate.model_dump(exclude_none=True)
            for candidate in candidates
        ]
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True, default_flow_style=False),
        encoding="utf-8",
    )


def _write_candidates_csv(candidates: list[OnboardingCandidate], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "company_name",
        "current_careers_url",
        "current_source_mode",
        "current_ats_type",
        "current_status_or_last_error",
        "candidate_official_website",
        "candidate_careers_url",
        "candidate_job_board_url",
        "detected_ats_type",
        "suggested_source_mode",
        "confidence",
        "needs_review",
        "reason",
        "suggested_action",
        "evidence",
        "approved",
        "sector",
        "category",
        "website_category",
        "ats_hint",
        "canada_hubs_notes",
        "role_families",
        "keywords",
        "priority",
        "monitoring_hint",
        "status",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()
        for candidate in candidates:
            row = candidate.model_dump()
            row["evidence"] = " | ".join(candidate.evidence)
            row["role_families"] = " | ".join(candidate.role_families)
            row["keywords"] = " | ".join(candidate.keywords)
            writer.writerow(row)


def write_candidates(candidates: list[OnboardingCandidate], output_path: Path) -> None:
    """Write candidates as YAML or CSV."""

    if output_path.suffix.lower() == ".csv":
        _write_candidates_csv(candidates, output_path)
        return
    _write_candidates_yaml(candidates, output_path)


def generate_candidates_from_input(
    *,
    input_path: Path,
    output_path: Path = DEFAULT_OUTPUT_PATH,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    starter_path: Path = DEFAULT_STARTER_PATH,
    reference_workbooks: tuple[Path, ...] = DEFAULT_REFERENCE_WORKBOOKS,
    live_discovery: bool = False,
    max_pages_per_company: int = 8,
    fetch_page: Any = None,
    robots_allowed: Any = None,
) -> list[OnboardingCandidate]:
    """Load onboarding input, generate candidates, and write output."""

    candidates = generate_candidates_from_records(
        load_company_inputs(input_path),
        companies_path=companies_path,
        starter_path=starter_path,
        reference_workbooks=reference_workbooks,
        live_discovery=live_discovery,
        max_pages_per_company=max_pages_per_company,
        fetch_page=fetch_page,
        robots_allowed=robots_allowed,
    )
    write_candidates(candidates, output_path)
    return candidates


def audit_large_company_list(
    *,
    input_path: Path = DEFAULT_LARGE_LIST_INPUT_PATH,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    starter_path: Path = DEFAULT_STARTER_PATH,
    readiness_output_path: Path = DEFAULT_READINESS_OUTPUT_PATH,
    candidates_output_path: Path = DEFAULT_LARGE_LIST_CANDIDATES_OUTPUT_PATH,
    report_output_path: Path = DEFAULT_LARGE_LIST_REPORT_PATH,
    needs_website_output_path: Path = DEFAULT_NEEDS_WEBSITE_OUTPUT_PATH,
    inspected_input_paths: tuple[Path, ...] = DEFAULT_REFERENCE_WORKBOOKS,
) -> dict[str, Any]:
    """Audit the 150-company spreadsheet for readiness without changing config."""

    spreadsheet_rows = load_spreadsheet_company_records(input_path)
    config_companies = [
        company
        for company in load_yaml_companies(companies_path)
        if isinstance(company, dict)
    ]
    starter_companies = [
        company
        for company in load_yaml_companies(starter_path)
        if isinstance(company, dict)
    ]
    config_index = _build_match_index(config_companies)
    starter_index = _build_match_index(starter_companies)

    readiness_rows: list[ReadinessAuditRow] = []
    generated_candidates: list[OnboardingCandidate] = []

    for record in spreadsheet_rows:
        existing_company, config_ambiguous = _resolve_index_match(
            config_index,
            record.company_name,
        )
        starter_company, starter_ambiguous = _resolve_index_match(
            starter_index,
            record.company_name,
        )
        best_url, best_url_source = _best_candidate_url(
            existing_company=existing_company,
            spreadsheet_record=record,
            starter_company=starter_company,
        )

        detected_ats_type: str | None = None
        suggested_source_mode: str | None = None
        if best_url:
            detected_ats_type, suggested_source_mode = _classify_candidate_source(
                company_name=record.company_name,
                careers_url=best_url,
                website_category=record.website_category,
                ats_hint=clean_text((existing_company or {}).get("ats_hint"))
                or record.website_category,
                current_source_mode=clean_text((existing_company or {}).get("source_mode")),
            )
        elif existing_company is not None:
            detected_ats_type, suggested_source_mode = _classify_candidate_source(
                company_name=record.company_name,
                careers_url=clean_text(existing_company.get("careers_url")),
                website_category=clean_text(existing_company.get("website_category")),
                ats_hint=clean_text(existing_company.get("ats_hint")),
                current_source_mode=clean_text(existing_company.get("source_mode")),
            )

        note_parts: list[str] = []
        if record.notes:
            note_parts.append(record.notes)
        if (
            record.spreadsheet_career_display_text
            and not record.spreadsheet_career_url_or_hyperlink
        ):
            note_parts.append(
                "spreadsheet careers cell has display text but no hyperlink target"
            )
        if config_ambiguous or starter_ambiguous:
            note_parts.append("multiple possible config/starter matches detected")

        readiness_status, recommended_next_action = _classify_readiness_status(
            existing_match=existing_company is not None,
            ambiguous_match=config_ambiguous or starter_ambiguous,
            best_url_source=best_url_source,
            source_mode=suggested_source_mode,
            display_text=record.spreadsheet_career_display_text,
        )
        usable_url_available = bool(best_url)

        readiness_row = ReadinessAuditRow(
            company_name=record.company_name,
            sector=record.sector,
            category=record.category,
            priority=record.priority,
            status=record.status,
            existing_config_match=existing_company is not None,
            existing_config_url=clean_text((existing_company or {}).get("careers_url")),
            spreadsheet_career_display_text=record.spreadsheet_career_display_text,
            spreadsheet_career_url_or_hyperlink=record.spreadsheet_career_url_or_hyperlink,
            starter_url_match=starter_company is not None,
            starter_url=clean_text((starter_company or {}).get("careers_url")),
            usable_url_available=usable_url_available,
            detected_ats_type=detected_ats_type,
            suggested_source_mode=suggested_source_mode or "needs_url",
            readiness_status=readiness_status,
            recommended_next_action=recommended_next_action,
            notes=" | ".join(note_parts) if note_parts else None,
            existing_source_mode=clean_text((existing_company or {}).get("source_mode")),
            existing_ats_hint=clean_text((existing_company or {}).get("ats_hint")),
            existing_status=clean_text((existing_company or {}).get("status")),
        )
        readiness_rows.append(readiness_row)

        if existing_company is not None or not usable_url_available:
            continue

        evidence = [
            f"source={best_url_source}",
            f"company_name={record.company_name}",
        ]
        if record.spreadsheet_career_display_text:
            evidence.append(
                f"display_text={record.spreadsheet_career_display_text}"
            )
        if starter_company is not None and best_url_source == "starter_url":
            evidence.append("starter_url_candidate=true")
        candidate = _build_candidate(
            company_name=record.company_name,
            source_record=_spreadsheet_record_to_source_record(record),
            careers_url=best_url,
            confidence="medium",
            reason=(
                "spreadsheet_hyperlink_candidate"
                if best_url_source == "spreadsheet_hyperlink"
                else "starter_career_url_match"
            ),
            evidence=evidence,
            suggested_action="review_and_apply",
        )
        generated_candidates.append(candidate)

    readiness_rows = sorted(
        readiness_rows,
        key=lambda item: _normalize_match_key(item.company_name),
    )
    generated_candidates = _dedupe_candidates(generated_candidates)
    _write_readiness_csv(readiness_rows, readiness_output_path)
    write_candidates(generated_candidates, candidates_output_path)
    _write_needs_website_csv(readiness_rows, needs_website_output_path)

    existing_inputs = tuple(path for path in inspected_input_paths if path.exists())
    summary = _build_large_list_report(
        inspected_inputs=existing_inputs or (input_path,),
        rows=readiness_rows,
        candidates=generated_candidates,
        output_path=report_output_path,
    )
    return {
        "rows": readiness_rows,
        "candidates": generated_candidates,
        "summary": summary,
        "readiness_output_path": readiness_output_path,
        "candidates_output_path": candidates_output_path,
        "report_output_path": report_output_path,
        "needs_website_output_path": needs_website_output_path,
    }


def load_candidate_file(input_path: Path) -> list[OnboardingCandidate]:
    """Load onboarding candidates from YAML or CSV."""

    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        with input_path.open("r", encoding="utf-8", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            items: list[OnboardingCandidate] = []
            for row in reader:
                row["needs_review"] = str(row.get("needs_review") or "").strip().lower() == "true"
                row["approved"] = str(row.get("approved") or "").strip().lower() == "true"
                row["evidence"] = [
                    item.strip()
                    for item in str(row.get("evidence") or "").split("|")
                    if item.strip()
                ]
                row["role_families"] = [
                    item.strip()
                    for item in str(row.get("role_families") or "").split("|")
                    if item.strip()
                ]
                row["keywords"] = [
                    item.strip()
                    for item in str(row.get("keywords") or "").split("|")
                    if item.strip()
                ]
                items.append(OnboardingCandidate.model_validate(row))
            return items

    payload = yaml.safe_load(input_path.read_text(encoding="utf-8")) or {}
    items = payload.get("candidates", payload if isinstance(payload, list) else [])
    return [OnboardingCandidate.model_validate(item) for item in items]


def _backup_file(path: Path) -> Path:
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    backup_path = path.with_suffix(path.suffix + f".bak.{timestamp}")
    shutil.copy2(path, backup_path)
    return backup_path


def _candidate_apply_url(candidate: OnboardingCandidate) -> str | None:
    for value in (
        candidate.candidate_job_board_url,
        candidate.candidate_careers_url,
    ):
        if is_real_url(value):
            return str(value).strip()
    return None


def _candidate_to_company_record(candidate: OnboardingCandidate) -> dict[str, Any] | None:
    careers_url = _candidate_apply_url(candidate)
    if not careers_url:
        return None
    if candidate.suggested_source_mode in RESTRICTED_SOURCE_MODES:
        return None
    if candidate.detected_ats_type == "restricted_board":
        return None
    if not candidate.sector or not candidate.category:
        return None

    record: dict[str, Any] = {
        "name": candidate.company_name,
        "sector": candidate.sector,
        "category": candidate.category,
        "careers_url": careers_url,
        "source_mode": candidate.suggested_source_mode,
        "status": candidate.status or "Watching",
        "role_families": candidate.role_families,
        "keywords": candidate.keywords,
    }
    optional_fields = {
        "website_category": candidate.website_category,
        "ats_hint": candidate.ats_hint or candidate.detected_ats_type,
        "canada_hubs_notes": candidate.canada_hubs_notes,
        "priority": candidate.priority,
        "monitoring_hint": candidate.monitoring_hint,
    }
    for field_name, value in optional_fields.items():
        if value is not None and value != []:
            record[field_name] = value
    return record


def apply_approved_candidates(
    *,
    input_path: Path,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    update_existing: bool = False,
) -> dict[str, Any]:
    """Apply only approved onboarding candidates to config/companies.yaml."""

    candidates = load_candidate_file(input_path)
    existing_payload = yaml.safe_load(companies_path.read_text(encoding="utf-8")) or {}
    companies = existing_payload.get("companies", [])
    if not isinstance(companies, list):
        raise ValueError(f"Expected 'companies' list in {companies_path}")

    company_index = {
        _normalize_company_name(company.get("name")): company
        for company in companies
        if isinstance(company, dict)
    }

    applied = 0
    skipped_unapproved = 0
    skipped_existing = 0
    skipped_missing_required = 0
    skipped_restricted = 0
    backup_path: Path | None = None

    for candidate in candidates:
        if not candidate.approved:
            skipped_unapproved += 1
            continue
        careers_url = _candidate_apply_url(candidate)
        if not careers_url:
            skipped_missing_required += 1
            continue
        if candidate.suggested_source_mode in RESTRICTED_SOURCE_MODES:
            skipped_restricted += 1
            continue
        if candidate.detected_ats_type == "restricted_board":
            skipped_restricted += 1
            continue

        key = _normalize_company_name(candidate.company_name)
        existing_company = company_index.get(key)
        if existing_company is not None and not update_existing:
            skipped_existing += 1
            continue

        if existing_company is not None:
            if backup_path is None:
                backup_path = _backup_file(companies_path)
            update_company_record_in_yaml(
                companies_path,
                company_name=candidate.company_name,
                updates={
                    "careers_url": careers_url,
                    "source_mode": candidate.suggested_source_mode,
                    "ats_hint": candidate.ats_hint or candidate.detected_ats_type,
                },
            )
            existing_company["careers_url"] = careers_url
            existing_company["source_mode"] = candidate.suggested_source_mode
            if candidate.ats_hint or candidate.detected_ats_type:
                existing_company["ats_hint"] = candidate.ats_hint or candidate.detected_ats_type
            applied += 1
            continue

        company_record = _candidate_to_company_record(candidate)
        if company_record is None:
            skipped_missing_required += 1
            continue

        if backup_path is None:
            backup_path = _backup_file(companies_path)
        companies.append(company_record)
        company_index[key] = company_record
        applied += 1

    if applied > 0:
        companies_path.write_text(
            yaml.safe_dump(
                {"companies": companies},
                sort_keys=False,
                allow_unicode=True,
                default_flow_style=False,
            ),
            encoding="utf-8",
        )

    return {
        "applied": applied,
        "skipped_unapproved": skipped_unapproved,
        "skipped_existing": skipped_existing,
        "skipped_missing_required": skipped_missing_required,
        "skipped_restricted": skipped_restricted,
        "backup_path": str(backup_path) if backup_path is not None else None,
    }


def _load_source_rows(db_path: Path) -> dict[str, dict[str, Any]]:
    if not db_path.exists():
        return {}
    connection = initialize_database(db_path)
    return {
        _normalize_company_name(row.get("company_name")): row
        for row in get_source_status_rows(connection)
    }


def _load_pending_interventions(db_path: Path) -> dict[str, dict[str, Any]]:
    if not db_path.exists():
        return {}
    connection = initialize_database(db_path)
    return {
        _normalize_company_name(row.get("company_name")): row
        for row in get_intervention_queue(connection)
    }


def _load_health_state(state_path: Path) -> dict[str, Any]:
    if not state_path.exists():
        return {"sources": {}}
    payload = json.loads(state_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        return {"sources": {}}
    payload.setdefault("sources", {})
    return payload


def _write_health_state(state_path: Path, payload: dict[str, Any]) -> None:
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _days_since(timestamp: str | None) -> int | None:
    if not timestamp:
        return None
    normalized = str(timestamp).strip().replace("Z", "+00:00").replace(" ", "T")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    now = datetime.now(UTC)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return max(0, (now - parsed.astimezone(UTC)).days)


def _is_due_for_health_check(
    state: dict[str, Any],
    *,
    company_name: str,
    min_days_between_checks: int,
    force: bool,
) -> bool:
    if force:
        return True
    source_state = state.get("sources", {}).get(_normalize_company_name(company_name), {})
    days = _days_since(source_state.get("last_health_check_at"))
    return days is None or days >= min_days_between_checks


def _is_problem_source(
    company: dict[str, Any],
    source_row: dict[str, Any] | None,
    pending_intervention: dict[str, Any] | None,
) -> tuple[bool, str]:
    source_mode = str(
        (source_row or {}).get("source_mode") or company.get("source_mode") or ""
    ).strip()
    if source_mode == "needs_url":
        return True, "needs_url"

    readiness = str((source_row or {}).get("readiness_label") or "").strip()
    remediation = str((source_row or {}).get("remediation_label") or "").strip()
    status = str((source_row or {}).get("status") or "").strip()
    last_error = str((source_row or {}).get("error") or "").strip().lower()
    consecutive_failures = int((source_row or {}).get("consecutive_failures", 0) or 0)
    pending_count = int((source_row or {}).get("pending_intervention_count", 0) or 0)

    if remediation in {"source_url_review", "source_error"}:
        return True, remediation
    if readiness in {"needs_url", "error"}:
        return True, readiness
    if any(marker in last_error for marker in HARD_ERROR_MARKERS):
        return True, "hard_error_marker"
    if consecutive_failures >= 2:
        return True, "repeated_failures"
    if status == "paused" and (pending_count >= 1 or pending_intervention is not None):
        return True, "pending_intervention"
    if pending_intervention is not None and int(
        pending_intervention.get("occurrence_count", 1) or 1
    ) >= 2:
        return True, "repeated_pending_intervention"
    return False, "not_problem"


def _build_refresh_seed_urls(
    company: dict[str, Any],
    source_row: dict[str, Any] | None,
) -> list[str]:
    current_url = (
        (source_row or {}).get("source_url")
        or company.get("careers_url")
    )
    normalized = _normalize_url(current_url)
    if not normalized:
        return []
    seeds = [normalized]
    root = _root_url(normalized)
    if root and root != normalized:
        seeds.append(root)
    deduped: list[str] = []
    for url in seeds:
        if url not in deduped:
            deduped.append(url)
    return deduped


def _refresh_input(company: dict[str, Any], source_row: dict[str, Any] | None) -> CompanyInput:
    return CompanyInput(
        company_name=str(company.get("name") or "").strip(),
        website_url=_root_url(
            (source_row or {}).get("source_url") or company.get("careers_url")
        ),
        careers_url=(source_row or {}).get("source_url") or company.get("careers_url"),
    )


def refresh_sources(
    *,
    output_path: Path = DEFAULT_REFRESH_OUTPUT_PATH,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    db_path: Path = DEFAULT_DATABASE_PATH,
    state_path: Path = DEFAULT_HEALTH_STATE_PATH,
    only_problem_sources: bool = False,
    company_name: str | None = None,
    force: bool = False,
    max_pages_per_company: int = 8,
    min_days_between_checks: int = 7,
    fetch_page: Any = None,
    robots_allowed: Any = None,
) -> list[OnboardingCandidate]:
    """Generate reviewable replacement candidates for configured sources."""

    config_companies = [
        company
        for company in load_yaml_companies(companies_path)
        if isinstance(company, dict)
    ]
    if company_name:
        normalized_name = _normalize_company_name(company_name)
        config_companies = [
            company
            for company in config_companies
            if _normalize_company_name(company.get("name")) == normalized_name
        ]

    source_rows = _load_source_rows(db_path)
    pending_interventions = _load_pending_interventions(db_path)
    state = _load_health_state(state_path)
    candidates: list[OnboardingCandidate] = []
    now = datetime.now(UTC).replace(microsecond=0).strftime("%Y-%m-%dT%H:%M:%SZ")

    for company in config_companies:
        name = str(company.get("name") or "").strip()
        if not name:
            continue
        if not _is_due_for_health_check(
            state,
            company_name=name,
            min_days_between_checks=min_days_between_checks,
            force=force,
        ):
            continue

        key = _normalize_company_name(name)
        source_row = source_rows.get(key)
        pending = pending_interventions.get(key)
        is_problem, problem_reason = _is_problem_source(company, source_row, pending)
        if only_problem_sources and not is_problem:
            continue

        refresh_input = _refresh_input(company, source_row)
        current_url = refresh_input.careers_url
        current_status_or_error = str(
            (source_row or {}).get("error")
            or (source_row or {}).get("status")
            or ""
        ).strip() or None
        findings = discover_live_candidates(
            company_name=name,
            start_urls=_build_refresh_seed_urls(company, source_row),
            max_pages=max_pages_per_company,
            fetch_page=fetch_page,
            robots_allowed=robots_allowed,
        )
        company_candidates: list[OnboardingCandidate] = []
        if findings:
            for finding in findings:
                candidate = _candidate_from_live_discovery(
                    company_input=refresh_input,
                    finding=finding,
                    source_record=company,
                    current_careers_url=current_url,
                    current_source_mode=str(
                        (source_row or {}).get("source_mode") or company.get("source_mode") or ""
                    ).strip()
                    or None,
                    current_ats_type=clean_text((source_row or {}).get("ats_type")),
                    current_status_or_last_error=current_status_or_error,
                )
                if candidate.suggested_action == "keep_existing" and is_problem:
                    candidate.suggested_action = "manual_review_required"
                company_candidates.append(candidate)
        else:
            company_candidates.append(
                _missing_candidate(
                    name,
                    official_website=refresh_input.website_url,
                    reason="source_refresh_failed",
                    evidence=[
                        "source=source_refresh",
                        f"current_url={current_url or '-'}",
                        f"problem_reason={problem_reason}",
                    ],
                )
            )

        candidates.extend(company_candidates)
        state.setdefault("sources", {})[key] = {
            "company_name": name,
            "source_url": current_url,
            "last_health_check_at": now,
            "last_health_status": "problem" if is_problem else "ok",
            "last_candidate_count": len(company_candidates),
            "last_error": current_status_or_error,
        }

    write_candidates(_dedupe_candidates(candidates), output_path)
    _write_health_state(state_path, state)
    return _dedupe_candidates(candidates)


def weekly_source_check(
    *,
    output_path: Path = DEFAULT_WEEKLY_OUTPUT_PATH,
    companies_path: Path = DEFAULT_COMPANIES_PATH,
    db_path: Path = DEFAULT_DATABASE_PATH,
    state_path: Path = DEFAULT_HEALTH_STATE_PATH,
    only_problem_sources: bool = True,
    force: bool = False,
    max_pages_per_company: int = 8,
    min_days_between_checks: int = 7,
    fetch_page: Any = None,
    robots_allowed: Any = None,
) -> list[OnboardingCandidate]:
    """Run the lightweight weekly failsafe refresh flow."""

    return refresh_sources(
        output_path=output_path,
        companies_path=companies_path,
        db_path=db_path,
        state_path=state_path,
        only_problem_sources=only_problem_sources,
        company_name=None,
        force=force,
        max_pages_per_company=max_pages_per_company,
        min_days_between_checks=min_days_between_checks,
        fetch_page=fetch_page,
        robots_allowed=robots_allowed,
    )


def print_generate_summary(candidates: list[OnboardingCandidate], output_path: Path) -> None:
    """Print a compact generation summary."""

    print(f"generated: {len(candidates)}")
    print(f"output: {output_path}")
    print(f"high_confidence: {sum(1 for item in candidates if item.confidence == 'high')}")
    print(f"needs_review: {sum(1 for item in candidates if item.needs_review)}")


def print_apply_summary(summary: dict[str, Any]) -> None:
    """Print a compact apply summary."""

    print(f"applied: {summary['applied']}")
    print(f"skipped_unapproved: {summary['skipped_unapproved']}")
    print(f"skipped_existing: {summary['skipped_existing']}")
    print(f"skipped_missing_required: {summary['skipped_missing_required']}")
    print(f"skipped_restricted: {summary['skipped_restricted']}")
    print(f"backup_path: {summary['backup_path'] or '-'}")


def build_parser() -> argparse.ArgumentParser:
    """Build the onboarding CLI parser."""

    parser = argparse.ArgumentParser(description="Review-first company/source onboarding")
    subparsers = parser.add_subparsers(dest="command", required=True)

    generate_parser = subparsers.add_parser(
        "generate",
        help="Generate reviewable onboarding candidates",
    )
    generate_parser.add_argument("--input", type=Path, required=True)
    generate_parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    generate_parser.add_argument("--companies", type=Path, default=DEFAULT_COMPANIES_PATH)
    generate_parser.add_argument("--starter", type=Path, default=DEFAULT_STARTER_PATH)
    generate_parser.add_argument("--live-discovery", action="store_true")
    generate_parser.add_argument("--max-pages-per-company", type=int, default=8)

    refresh_parser = subparsers.add_parser(
        "refresh-sources",
        help="Generate reviewable replacement candidates for existing sources",
    )
    refresh_parser.add_argument("--output", type=Path, default=DEFAULT_REFRESH_OUTPUT_PATH)
    refresh_parser.add_argument("--companies", type=Path, default=DEFAULT_COMPANIES_PATH)
    refresh_parser.add_argument("--db-path", type=Path, default=DEFAULT_DATABASE_PATH)
    refresh_parser.add_argument("--state-path", type=Path, default=DEFAULT_HEALTH_STATE_PATH)
    refresh_parser.add_argument("--only-problem-sources", action="store_true")
    refresh_parser.add_argument("--company", type=str, default=None)
    refresh_parser.add_argument("--force", action="store_true")
    refresh_parser.add_argument("--max-pages-per-company", type=int, default=8)
    refresh_parser.add_argument("--min-days-between-checks", type=int, default=7)

    weekly_parser = subparsers.add_parser(
        "weekly-source-check",
        help="Run the lightweight weekly source health check",
    )
    weekly_parser.add_argument("--output", type=Path, default=DEFAULT_WEEKLY_OUTPUT_PATH)
    weekly_parser.add_argument("--companies", type=Path, default=DEFAULT_COMPANIES_PATH)
    weekly_parser.add_argument("--db-path", type=Path, default=DEFAULT_DATABASE_PATH)
    weekly_parser.add_argument("--state-path", type=Path, default=DEFAULT_HEALTH_STATE_PATH)
    weekly_parser.add_argument("--only-problem-sources", action="store_true")
    weekly_parser.add_argument("--force", action="store_true")
    weekly_parser.add_argument("--max-pages-per-company", type=int, default=8)
    weekly_parser.add_argument("--min-days-between-checks", type=int, default=7)

    apply_parser = subparsers.add_parser(
        "apply",
        help="Apply only approved onboarding candidates",
    )
    apply_parser.add_argument("--input", type=Path, required=True)
    apply_parser.add_argument("--companies", type=Path, default=DEFAULT_COMPANIES_PATH)
    apply_parser.add_argument("--update-existing", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    """CLI entrypoint."""

    args = build_parser().parse_args(argv)
    if args.command == "generate":
        candidates = generate_candidates_from_input(
            input_path=args.input,
            output_path=args.output,
            companies_path=args.companies,
            starter_path=args.starter,
            live_discovery=args.live_discovery,
            max_pages_per_company=args.max_pages_per_company,
        )
        print_generate_summary(candidates, args.output)
        return 0
    if args.command == "refresh-sources":
        candidates = refresh_sources(
            output_path=args.output,
            companies_path=args.companies,
            db_path=args.db_path,
            state_path=args.state_path,
            only_problem_sources=args.only_problem_sources,
            company_name=args.company,
            force=args.force,
            max_pages_per_company=args.max_pages_per_company,
            min_days_between_checks=args.min_days_between_checks,
        )
        print_generate_summary(candidates, args.output)
        return 0
    if args.command == "weekly-source-check":
        candidates = weekly_source_check(
            output_path=args.output,
            companies_path=args.companies,
            db_path=args.db_path,
            state_path=args.state_path,
            only_problem_sources=args.only_problem_sources or True,
            force=args.force,
            max_pages_per_company=args.max_pages_per_company,
            min_days_between_checks=args.min_days_between_checks,
        )
        print_generate_summary(candidates, args.output)
        return 0
    if args.command == "apply":
        summary = apply_approved_candidates(
            input_path=args.input,
            companies_path=args.companies,
            update_existing=args.update_existing,
        )
        print_apply_summary(summary)
        return 0
    raise ValueError(f"Unsupported onboarding command: {args.command}")


if __name__ == "__main__":
    raise SystemExit(main())
